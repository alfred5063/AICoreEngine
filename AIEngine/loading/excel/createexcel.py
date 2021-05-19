#!/usr/bin/python
# FINAL SCRIPT updated as of 15th April 2021

# Declare Python libraries needed for this script
import pandas as pd
from pandas import ExcelWriter
import os
from directory.get_filename_from_path import get_file_name
from utils.audit_trail import audit_log
from utils.logging import logging
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime as datetime
import numpy as np

#  """ Purpose: This function is use for create excel file
#      Author: Yeoh Eik Den
#      Created on: 17 May 2019"""

def create_excel(properties, df, base, sheetname ='sheet1', filename=None, **kwargs):
  try:
    audit_log('create_excel', 'export dataframe as excel file', base)
    head, tail = get_file_name(properties.destination)
    if filename==None: 
      writer = ExcelWriter(properties.destination, engine='openpyxl')
    else:
      writer = ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, sheetname, index=False)
    writer.save()
    writer.close()
    print("Excel file created:  %s" % tail)
  except Exception as error:
    logging('create_excel', error, base)
    raise error

def create_excel_sheet(columns_index,columns_text_index, df, sheetname, writer, move_to=None):
  workbook  = writer.book
  if 'MARC' in workbook.sheetnames:
    marc_workbook = workbook['MARC']
    marc_workbook.title = 'raw_source'

  df.to_excel(writer, sheetname, index=False, na_rep=np.nan)
  workbook = excel_reformat_date(workbook, sheetname, df, columns_index,columns_text_index, move_to)

def format_date(df, columns_index):
   for i in columns_index:
     try:
       print(df.iloc[:, i])
       df.iloc[:,i-1] = df.iloc[:,i-1].apply(lambda x: datetime.datetime.strptime(x, "%d/%m/%Y") if (x!='') else np.nan)
     except:
       try:
         df.iloc[:,i-1] = df.iloc[:,i-1].apply(lambda x: datetime.datetime.strptime(x, "%Y-%m-%d %H:%M:%S") if (x!='') else np.nan)
       except Exception as error:
         pass
       #print(error)
       #print('converting date format to %Y-%m-%d %H:%M:%S')
       #df.iloc[:,i-1] = df.iloc[:,i-1].apply(lambda x: datetime.datetime.strptime(x, "%Y-%m-%d %H:%M:%S") if (x!='') else np.nan)
   return df

def create_excel_dm(properties, df, base, columns_index, columns_text_index, sheetname ='result', **kwargs):
    
    audit_log('create_excel', 'export dataframe as excel file', base)
    head, tail = get_file_name(properties.source)
    writer = ExcelWriter(head + "\\result_"+tail, engine='openpyxl')

    try:
      writer.book = load_workbook(head + "\\result_"+tail)
    except:
      writer.book = load_workbook(head + "\\"+tail)

    df = format_date(df, columns_index)

    create_excel_sheet(columns_index,columns_text_index, df, sheetname, writer)

    df_unsuccessful = df[df['record_status']=='Unsuccessful']
    df_successful = df[df['record_status']=='Successful']
    
    #DELETE Sheet
    df_unsuccessful_cancel_delete = df_unsuccessful[(df_unsuccessful['action'].str.contains("cancel_delete")) & (df_unsuccessful['import type'] == 'X')]
    df_successful_cancel_delete = df_successful[(df_successful['action'].str.contains("import")) & (df_successful['import type']=='X')]
    df_cancel_delete= pd.concat([df_unsuccessful_cancel_delete, df_successful_cancel_delete])

    #MARC Sheet
    df_import = df_successful[df_successful['action'].str.contains("import")]

    #RPA Sheet
    df_rpa = df_successful[~df_successful['action'].str.contains("import")]

    try:
      df_unsuccessful_without_delete = df_unsuccessful[(~df_unsuccessful['action'].str.contains("cancel_delete"))]
      #create result populated by RPA
      create_excel_sheet(columns_index,columns_text_index, df_unsuccessful_without_delete, 'MANUAL', writer, move_to=0)
    except Exception as error:
      print('no record in manual sheet')
      logging('create_excel', error, base)

    try:
      #create result populated by RPA
      df_rpa = removed_columns_error(df_rpa)
      create_excel_sheet(columns_index,columns_text_index, df_rpa, 'RPA', writer, move_to=0)
    except Exception as error:
      print('no record in RPA sheet')
      logging('create_excel', error, base)

    try:
      #create delete sheet
      df_cancel_delete = removed_columns(df_cancel_delete)
      create_excel_sheet(columns_index,columns_text_index, df_cancel_delete, 'DEL', writer, move_to=0)
    except Exception as error:
      print('no record in DEL sheet')
      logging('create_excel', error, base)

    try:
      #create marc
      df_import = removed_columns(df_import)
      create_excel_sheet(columns_index,columns_text_index, df_import, 'MARC', writer, move_to=0)
    except Exception as error:
      print('no record in RPA sheet')
      logging('create_excel', error, base)

    writer.save()
    writer.close()
    print("Excel file created:  %s" % tail)


def removed_columns(df):
  #action_code	fields_update	action	search_criteria	error	record_status
  result = df.drop(['action_code','fields_update',	'action',	'search_criteria', 'error', 'record_status'], axis=1)
  return result

def removed_columns_error(df):
  #action_code	fields_update	action	search_criteria	error	record_status
  result = df.drop(['error'], axis=1)
  return result

def excel_reformat_date(workbook, sheetname, df, columns_index,columns_text_index, move_to=None):
   worksheet = workbook[sheetname]
   #create result populated by RPA
   
   if move_to != None:
    move_sheet(workbook, workbook.get_index(worksheet), move_to)
   row, column = df.shape
   columns = columns_index
   columns_text = columns_text_index
   rows = row + 2
   
   for y in columns:
     for x in range(1,rows):
       worksheet.cell(row=x,column=y).number_format = 'dd/mm/yyyy'

   return workbook

def move_sheet(wb, from_loc=None, to_loc=None):
    sheets=wb._sheets

    # if no from_loc given, assume last sheet
    if from_loc is None:
        from_loc = len(sheets) - 1

    #if no to_loc given, assume first
    if to_loc is None:
        to_loc = 0

    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)

def export_excel_split_duplicate(properties, non_duplicated, duplicated, not_found):
  head, tail = get_file_name(properties.destination)
  writer = pd.ExcelWriter(properties.destination)
  non_duplicated.to_excel(writer, sheet_name='non-duplicatd', index=False)
  duplicated.to_excel(writer, sheet_name='duplicated', index=False)
  not_found.to_excel(writer, sheet_name='not found', index=False)
  writer.save()
  writer.close()
  print("Excel file created:  %s" % tail)

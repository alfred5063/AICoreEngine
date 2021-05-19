#!/usr/bin/python
# FINAL SCRIPT updated as of 22nd June 2020
# Workflow - CBA/ADMISSION

# Declare Python libraries needed for this script
from transformation.adm_bdx_manipulation import build_rb,read_BdxL,count_BdxL,count_BdxL_Colomn
import openpyxl
import pandas as pd
from xlrd import *
from utils.audit_trail import audit_log
import json
from automagic.marc_adm import *
from transformation.adm_connect_db import *
from openpyxl.styles import Font,Border, Side
import datetime as d
from loading.excel.checkExcelLoading import is_locked
import copy,json
from datetime import datetime
import datetime
import time
from transformation.adm_fill_bdx import Check_post_or_type
from datetime import datetime as dt
from connector.connector import MySqlConnector


def search_excel(keyword,sheet, workbook):
  ws = workbook.worksheets[sheet]
  for row in ws:
    for cell in row:
      if cell.value == keyword:
        cell_no = [cell.column,cell.row]
        return cell_no
        break

def search_last_row(cell_no, CMfile, sheet, workbook, keyword):
  #Search the latest row to append data in excel file from dataframe
  df = pd.read_excel(CMfile, header = cell_no[1]-1, sheet_name=sheet)
  main_df = df.ix[:,keyword]
  main_df = main_df.dropna()
  main_df.last_valid_index()
  search_excel(main_df[main_df.first_valid_index()],sheet, workbook)[1]
  if(main_df.first_valid_index()==None):
    return cell_no[1]+2
  else:
    return search_excel(main_df[main_df.first_valid_index()],sheet, workbook)[1]+main_df.last_valid_index()-main_df.first_valid_index()

def bdx_to_CM(BDFile,CMFile,db_bdx_invoice_type,adm_base,path=None):
  if path == None:
    path = CMFile

  if is_locked(CMFile):
    audit_log("CMFILE cannot be opened", "Completed...", adm_base)
    return None
  else:
    audit_log("CMFILE can be opened", "Completed...", adm_base)
    wb = openpyxl.load_workbook(CMFile)

  All_Sheet=wb.sheetnames
  Lower_Case_Sheet_Name=[]
  for i in All_Sheet:
    Lower_Case_Sheet_Name.append(i.lower())
  Find_sheet=[]
  db_bdx_invoice_type=db_bdx_invoice_type.lower()
  type_list=db_bdx_invoice_type.split(" ")
  counter=0
  for sheet in Lower_Case_Sheet_Name:
    for i in range(len(type_list)):
      if type_list[i] in sheet:
        if i+1==len(type_list):
          Find_sheet.append(All_Sheet[counter])
    counter+=1
  if len(Find_sheet) is 2:
    if len(type_list) ==1:
      status=True # True is mean B2B or Post only
    else:
      status=False#post b2b
    if status:# lesser word
      for i in range(len(Find_sheet)):
        if (len(Find_sheet[0].split())) > (len(Find_sheet[1].split())):
          Find_sheet[0]=Find_sheet[1]
          continue
        else:
          continue
    else:
      for sheet in Find_sheet:
        if len(Find_sheet[0].split())>len(Find_sheet[1].split()):
          continue
        else:
          Find_sheet[0]=Find_sheet[1]
          continue
  elif len(Find_sheet)==1:
    audit_log("{} is the invoice type".format(Find_sheet[0]), "Completed...", adm_base)
  else:
    audit_log("CM invoice type might be wrong, place check again", "Completed...", adm_base)


  Sheet_Index=All_Sheet.index(Find_sheet[0])
  print("type ",Find_sheet[0])
  print("index ",Sheet_Index)
  if len(Find_sheet)==0:
    Find_sheet.append(All_Sheet[0])

  ws=wb[Find_sheet[0]]
  starting_row=search_last_row(['D',5], CMFile, Sheet_Index, wb, "Disbursement Listing")+1+1
  ending_row=search_last_row(['D',5], CMFile,Sheet_Index , wb, "Disbursement Listing")+1+1+count_BdxL(BDFile)
  if starting_row==ending_row:
    ending_row+=1
  num_of_colomn=count_BdxL_Colomn(BDFile)
  ft1 = Font(name='Trebuchet MS', size=8)
  thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
  skiprow_sheet=0
  for i in range(1,100):
    if ws.cell(row=i, column=2).value=="Adm. No":
      skiprow_sheet=i
      break
  print(skiprow_sheet,"is skiprows")

  rb=open_workbook(BDFile,formatting_info=True)
  readerSheet = rb.sheet_by_index(0)
  bdx_lisitng=readerSheet.cell(12,1)
  date=readerSheet.cell(13,1)
  for i in range(starting_row,ending_row):
    ws.cell(row=i, column=16).value=date.value
    ws.cell(row=i, column=16).font=ft1
    ws.cell(row=i, column=16).border = thin_border
    ws.cell(row=i, column=4).value=bdx_lisitng.value
    ws.cell(row=i, column=4).font=ft1
    ws.cell(row=i, column=4).border = thin_border

  dicbdx={2:"File No",
          3:"File No",
          5:"Disbursement No",
          6:"Patient Name",
          7:"Policy No",
          8:"Hosp Bill No",
          9:"Admission",
          10:"Discharge",
          11:"Hospital",
          13:"Insurance",
          14:"Excess",
          15:"Total Bill",
    }

  df=read_BdxL(BDFile)
  json_list=[]
  data=df.to_json(orient='index')
  admNoList=[]
  for column in range(num_of_colomn):
    for DF_coloum in df.T.index:
      list_to_update_in_excel=[]
      counter=0
      if DF_coloum == "No":
        continue
      if DF_coloum == "File No":
        list_to_update_in_excel.append(2)
        list_to_update_in_excel.append(3)
      if DF_coloum == "Hosp Bill No":
        list_to_update_in_excel.append(8)
      if DF_coloum == "Disbursement No":
        list_to_update_in_excel.append(5)
      if DF_coloum == "Patient Name":
        list_to_update_in_excel.append(6)
      if DF_coloum == "Staff ID":
        continue
      if DF_coloum == "Policy No":
        list_to_update_in_excel.append(7)
      if DF_coloum == "Admission":
        list_to_update_in_excel.append(9)
      if DF_coloum=="Discharge":
        list_to_update_in_excel.append(10)
      if DF_coloum=="Hospital":
        list_to_update_in_excel.append(11)
      if DF_coloum=="Insurance":
        list_to_update_in_excel.append(13)
      if DF_coloum=="Excess":
        list_to_update_in_excel.append(14)
      if DF_coloum=="Total Bill":
        list_to_update_in_excel.append(15)
      if DF_coloum=="Remarks":
        continue
      for i in list_to_update_in_excel:
        list_to_update_df=list(df.T.loc[dicbdx[i]])
        if i==3:
          admNoList=list(df.T.loc[dicbdx[i]]).copy()

        for row_no in range(starting_row,ending_row):
          ws.cell(row=row_no, column=i).value = list_to_update_df[counter]
          ws.cell(row=row_no, column=i).font=ft1
          ws.cell(row=row_no, column=i).border = thin_border
          counter+=1
        list_to_update_df.clear()
        counter=0

  
  if is_locked(CMFile):
    audit_log("CMfile is locked", "Completed...", adm_base)
  else:
    wb.save(path)
    audit_log("CMfile is saved", "Completed...", adm_base)


def bdx_to_CM_database(BDFile, bdx_type, cliend_id, adm_base):
  email = adm_base.email
  curyear = time.strftime("%y", time.localtime())
  current_year = dt.now().year
  named_tuple = time.localtime()
  current_date = time.strftime("%Y-%m-%d", named_tuple)

  rb = open_workbook(BDFile, formatting_info = True)
  readerSheet = rb.sheet_by_index(0)
  for i in range(20):
    if "Bord No" in str(readerSheet.cell(i,0).value):
      bdx_lisitng = readerSheet.cell(i,1).value
    if "Submission" in str(readerSheet.cell(i,0).value):
      date = readerSheet.cell(i,1).value
      date = datetime.datetime.strptime(date, "%d/%m/%Y").strftime("%Y-%m-%d")

  df = read_BdxL(BDFile)
  Disbursement_No = df["Disbursement No"][0]
  num_of_colomn = count_BdxL_Colomn(BDFile)
  total_bord_row = count_BdxL(BDFile)
  bord_dict = json.loads(df.to_json(orient = 'index'))
  cm_json = {
            "SAGE DOC ID": "", "Adm. No": "", "File No.": "", "Disbursement Listing": "", "Disbursement Claims": "", "Member's Name": "",
            "Policy No": "", "Plan Type": "", "Adm. Date": "", "Dis. Date": "", "Hospital": "", "Diagnosis": "", "Insurance": "", "Patient": "",
            "Discrepancy": "", "Total": "", "Submission Date": "", "OB Received Date": "", "OB Registered Date": "", "CSU Received Date": "",
            "Primary GL No": "", "Remarks": "", "Payment Due Date": "", "Aging (days)": "", "SCAN (TAT)": "", "CSU (TAT)": "", "DB BANK CODE": "",
            "Under / (Over) Claimed fr Client (AA-J-AM+AP)": "", "Cheque DetailsBank-in Date": "", "Cheque DetailsChq.No.": "",
            "Cheque DetailsAmount (RM)": "", "Hospital Bills DetailsDate": "", "Hospital Bills DetailsBill No.": "", "Hospital Bills DetailsAmount (RM)": "",
            "Under / (Over) Reimbursed by Client(AB-Y-AN+AQ)": "", "UP/OP CASES": "", "AAN Settlement DetailsDate": "", "AAN Settlement DetailsChq.No.": "",
            "AAN Settlement DetailsAmount (RM)": "", "Under / (Over) Paid to Hosp (AB - AG)": "", "Designated  / Borrowing Bank AccountOCBC 16": "",
            "Designated  / Borrowing Bank AccountBIMB 1": "", "Designated  / Borrowing Bank AccountDBMY29": "",
            "Designated  / Borrowing Bank AccountDBMY29.1": "", "Designated  / Borrowing Bank AccountDBMY29.2": "",
            "Designated  / Borrowing Bank AccountDBMY29.3": "", "Designated  / Borrowing Bank AccountDBMY29.4": "", "TPA Billing DetailsDate": "",
            "TPA Billing DetailsInv. No.": "", "TPA Billing DetailsAmount (RM)": "", "Reimbursement DetailsDate": "", "Reimbursement DetailsChq. No.": "",
            "Reimbursement DetailsAmount (RM)": "", "Adjustment DetailsDate": "", "Adjustment DetailsDoc. No.": "", "Adjustment DetailsAmount (RM)": "",
            "Hospital OR DetailsOR Date": "", "Hospital OR DetailsOR No.": "", "Hospital OR DetailsAmount (RM)": "", "Hospital OR DetailsDate of OR Rec'd": "",
            "Hospital OR DetailsSubm. Date": "", "Tax Invoice DetailsDate": "", "Tax Invoice DetailsInv No": "", "Tax Invoice DetailsCase Fee": "",
            "Tax Invoice DetailsGST Amount": "", "Tax Invoice DetailsTotal": ""
            }
  

  for row in range(total_bord_row):

    dict_to_update = copy.deepcopy(cm_json)

    if '-' in df.iloc[row]['File No']:
      filenum = df.iloc[row]['File No'].split('-')
      caseid = int(filenum[1])
    else:
      caseid = int(df.iloc[row]['File No'])

    type = Check_post_or_type(caseid)
    parameters = [str(caseid),]
    conn_mysql = MySqlConnector()
    cur_mysql = conn_mysql.cursor()

    if type == 'POST':
      stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
      for z in cur_mysql.stored_results():
        results = z.fetchall()
      results_cln = pd.DataFrame(results)

      stored_proc = cur_mysql.callproc('query_marc_for_outpatient_validity', parameters)
      for z in cur_mysql.stored_results():
        results = z.fetchall()
      fetched_result = pd.DataFrame(results)
      results_cln = pd.DataFrame(results_cln)
      fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg','CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName',  'CaseMmClientName',
                                'CaseMmInsurerName', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate', 'CasePpId']
      if results_cln.empty != True:
        results_cln.columns = ['Sub Case ID', 'Client Listing Number']
        fetched_result['CaseAccChqNum'] = results_cln['Client Listing Number']
      else:
        fetched_result['CaseAccChqNum'] = "N/A"

    elif type == 'CASHLESS':

      stored_proc = cur_mysql.callproc('query_marc_for_inpatient_validity', parameters)
      for z in cur_mysql.stored_results():
        results = z.fetchall()
      fetched_result = pd.DataFrame(results)
      fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName', 'CaseMmClientName',
                                'CaseMmInsurerName', 'CaseAccChqNum', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate']
      fetched_result['CasePpId'] = "0"

    else:

      stored_proc = cur_mysql.callproc('query_marc_for_inpatient_validity', parameters)
      for z in cur_mysql.stored_results():
        results = z.fetchall()
      fetched_result = pd.DataFrame(results)
      fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName', 'CaseMmClientName',
                                'CaseMmInsurerName', 'CaseAccChqNum', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate']
      fetched_result['CasePpId'] = "0"

    for column in range(num_of_colomn):

      if list(bord_dict[str(0)].keys())[column] == "File No":
        dict_to_update["Adm. No"] = bord_dict[str(row)]["File No"]
        dict_to_update["SAGE DOC ID"] = bord_dict[str(row)]["File No"]
        dict_to_update["File No."] = bord_dict[str(row)]["File No"]

      if list(bord_dict[str(0)].keys())[column]=="Hosp Bill No":
        dict_to_update["Bill No"]=bord_dict[str(row)]["Hosp Bill No"]

      if list(bord_dict[str(0)].keys())[column]=="Disbursement No":
        dict_to_update["Disbursement Claims"]=bord_dict[str(row)]["Disbursement No"]

      if list(bord_dict[str(0)].keys())[column]=="Patient Name":
        dict_to_update["Member's Name"]=bord_dict[str(row)]["Patient Name"]

      if list(bord_dict[str(0)].keys())[column]=="Policy No":
        dict_to_update["Policy No"]=bord_dict[str(row)]["Policy No"]

      if list(bord_dict[str(0)].keys())[column]=="Admission":
        dict_to_update["Adm. Date"]=bord_dict[str(row)]["Admission"]

      if list(bord_dict[str(0)].keys())[column]=="File No":
        dict_to_update["Adm. No"]=bord_dict[str(row)]["File No"]

      if list(bord_dict[str(0)].keys())[column]=="Hospital":
        dict_to_update["Hospital"]=bord_dict[str(row)]["Hospital"]

      if list(bord_dict[str(0)].keys())[column]=="Insurance":
        dict_to_update["Insurance"]=bord_dict[str(row)]["Insurance"]

      if list(bord_dict[str(0)].keys())[column]=="Excess":
        dict_to_update["Patient"]=bord_dict[str(row)]["Excess"]

      if list(bord_dict[str(0)].keys())[column]=="Total Bill":
        dict_to_update["Total"]=bord_dict[str(row)]["Total Bill"]

      # This one from CML
      if list(bord_dict[str(0)].keys())[column]=="Remarks":
        dict_to_update["Remarks"]=bord_dict[str(row)]["Remarks"]

      bdx_lisitng = bdx_lisitng.split("/")
      bdx_lisitng = str(bdx_lisitng[0].zfill(5)) + "/" + str(bdx_lisitng[1])
      dict_to_update["Disbursement Listing"] = bdx_lisitng
      dict_to_update["Submission Date"] = date

      if fetched_result['inptCaseDiscDate'][int(row)] == None:
        dict_to_update["Dis. Date"] = ''
      else:
        dict_to_update["Dis. Date"]=fetched_result['inptCaseDiscDate'][int(row)].strftime("%Y-%m-%d")

      if fetched_result['inptCaseBillReceivedDate'][int(row)] == None:
        dict_to_update["OB Received Date"] = fetched_result['inptCaseBillRegDate'][int(row)].strftime("%Y-%m-%d")
      else:
        dict_to_update["OB Received Date"] = fetched_result['inptCaseBillReceivedDate'][int(row)].strftime("%Y-%m-%d")

      if fetched_result['inptCaseBillRegDate'][int(row)] == None:
        dict_to_update["OB Registered Date"] = ''
      else:
        dict_to_update["OB Registered Date"] = fetched_result['inptCaseBillRegDate'][int(row)].strftime("%Y-%m-%d")

      dict_to_update["CSU Received Date"] = current_date

    try:
      update_cm_to_database(dict_to_update, cliend_id, bdx_type, email)
    except:
      update_new_cm_to_database(dict_to_update,cliend_id,bdx_type,email)

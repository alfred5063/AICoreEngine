#!/usr/bin/python
# FINAL SCRIPT updated as of 15th April 2021
# Workflow - Finance SOA
# Version 1

# Declare Python libraries needed for this script
import openpyxl
import time
import pandas as pd
import json as json
import numpy as np
import re
import shutil
import sys
import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime as dt
from xlrd import open_workbook
from pathlib import Path
from os import listdir
from os.path import isfile, join
from utils.audit_trail import audit_log
from utils.logging import logging
from difflib import SequenceMatcher
from transformation.reqdca_excel_manipulation import *
from loading.excel.checkExcelLoading import is_locked

# Client Master File mapping List('MCIS Zurich', 'MSIG', 'HLA-LIFE', 'ALLIANZ GENERAL')('PACIFIC', 'AXA AFFIN')
cmf_list = ['AM ASSURANCE', 'BERJAYA SOMPO', 'GIBRALTAR BSN LIFE', 'GREAT EASTERN GENERAL', 'PROGRESSIVE', 'RHB', 'TISB', 'SYARIKAT TAKAFUL MALAYSIA BERHAD', 'TUNE INSURANCE', 'SUNLIFE', 'SS2']
#cmf_list = ['HLA-LIFE']

# Get current year
current_year = dt.now().year
last_year = current_year - 1

dirName = r'\\fs1\finance\1-CREDIT CONTROL\2-MASTER & OPS INVOICES\MASTER'
def getListOfFiles(dirName):
    # create a list of file and sub directories 
    # names in the given directory
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory 
        if os.path.isdir(fullPath):
           allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)
    return allFiles

list_df = pd.DataFrame(getListOfFiles(dirName))
list_df.columns = ['Path']

def edited_cmf(soa_df, base):

    # Create empty dataframe to store data 
    final_df = pd.DataFrame()
    soa_df['Remarks'] = soa_df['Remarks'].replace('nan', np.NaN, regex = True)
    soa_main_df = soa_df[pd.notnull(soa_df['GL Ref']) & pd.isnull(soa_df['Remarks'])]
    soa_main_df.reset_index(drop = True, inplace = True)
    a = 0
    for a in range(len(cmf_list)):
      #path = dirName + '\\' + cmf_list[x]
      extracted_df = list_df[list_df['Path'].str.contains(cmf_list[a])]
      extracted_df = list_df[list_df['Path'].str.contains(cmf_list[a])]
      extracted_df = extracted_df[extracted_df['Path'].str.contains(str(current_year))]
      extracted_ls = list(extracted_df['Path'])
      if extracted_ls == []:
        extracted_df = list_df[list_df['Path'].str.contains(cmf_list[a])]
        extracted_df = extracted_df[extracted_df['Path'].str.contains(str(last_year))]
        extracted_ls = list(extracted_df['Path'])
      else:
        pass
      #To remove the file contains of ~$
      b = 0
      for b in range(len(extracted_ls)):
        if '~$' in extracted_ls[b]:
          extracted_ls.remove(extracted_ls[b])
        else:
          pass
        continue

      # Process each file
      c = 0
      for c in range(len(extracted_ls)):
        # An empty dataframe to store data
        right_df = pd.DataFrame(columns = ['Adm. No', 'Insurance', 'Submission Date', 'Bank-in Date', 'Amount (RM)'])

        # Set the file location
        filelocation = extracted_ls[c]

        # Provide the name of the range of the header to be extracted
        first_header = 'Adm. No'
        second_header = 'Bank-in Date'
        last_header = 'Date of OR Rec\'d'
        filestatus = is_locked(filelocation)
        if filestatus == False:
          wb = load_workbook(filelocation, read_only = False, keep_vba = True, data_only = False, keep_links = True)
          sheets = wb.sheetnames

          d = 0
          for d in range(len(sheets)):
            if sheets[d] != 'DCA' and sheets[d] != 'NO BORD':
              # Get the cell position of the first and last header
              start_cell = search_excel(first_header, d, wb)
              end_cell = search_excel(last_header, d, wb)

              # Read Workbook as dataframe
              if start_cell[1] == 1:
                searched_df = pd.read_excel(filelocation, sheet_name = sheets[d])
              else:
                searched_df = pd.read_excel(filelocation, sheet_name = sheets[d], skiprows = start_cell[1]-1)

              # Store the correct column index in a list
              columns_ls = list(searched_df.columns)
              column_index_ls = []
              column_index_ls.append(columns_ls.index('Adm. No'))
              column_index_ls.append(columns_ls.index('Insurance'))
              column_index_ls.append(columns_ls.index('Submission Date'))
              bankinDate_number = columns_ls.index('Cheque Details')
              column_index_ls.append(columns_ls.index('Cheque Details'))
              column_index_ls.append(bankinDate_number + 2)

              # Store the correct column data in a df
              e = 0
              for e in range(len(column_index_ls)):
                right_df[right_df.columns[e]] = searched_df[searched_df.columns[column_index_ls[e]]]

              # Clean up
              right_df = right_df[pd.notnull(right_df['Adm. No'])]
              right_df.reset_index(drop = True, inplace = True)

              right_adm_ls = []
              adm_no_ls = []
              gl_no_ls = list(soa_main_df['GL Ref'])
              validated_df = pd.DataFrame([])
              if right_df.empty == False:
                f = 0
                for f in range(len(right_df['Adm. No'])):
                  if '-' in str(right_df['Adm. No'][f]):
                    adm_no_ls.append(str(right_df['Adm. No'][f]))
                    if str(right_df['Adm. No'][f]) in gl_no_ls:
                      right_adm_ls.append(right_df['Adm. No'][f])
                    else:
                      pass
                  elif '.' in str(right_df['Adm. No'][f]):
                    right_df.loc[f, 'Adm. No'] = str(right_df['Adm. No'][f])[:str(right_df['Adm. No'][f]).index('.')]
                    gl_ref = right_df.loc[f, 'Adm. No']
                    adm_no_ls.append(right_df['Adm. No'][f])
                    if str(gl_ref) in gl_no_ls:
                      right_adm_ls.append(right_df['Adm. No'][f])
                    else:
                      pass
                  else:
                    adm_no_ls.append(str(right_df['Adm. No'][f]))
                    if str(right_df['Adm. No'][f]) in gl_no_ls:
                      right_adm_ls.append(right_df['Adm. No'][f])
                    else:
                      pass
              else:
                pass

              validate_ls = []
              if right_adm_ls != []:
                results = pd.DataFrame([])
                g = 0
                for g in range(len(right_adm_ls)):
                  row_index = int(adm_no_ls.index(right_adm_ls[g]))
                  validated_df = validated_df.append(right_df.iloc[row_index])
                  validated_df.reset_index(drop=True, inplace=True)
                  validate_ls = list(validated_df['Adm. No'])
                  validate_index = int(validate_ls.index(right_adm_ls[g]))
                  results = results.append(soa_main_df[soa_main_df['GL Ref'].astype(str).str.contains("%s" % validated_df['Adm. No'][validate_index])])

                results.reset_index(drop=True, inplace=True)
                results['Insurance'] = validated_df['Insurance']
                results['Submission Date'] = validated_df['Submission Date']
                results['Bank-in Date'] = validated_df['Bank-in Date']
                results['Amount (RM)'] = validated_df['Amount (RM)']

                final_df = final_df.append(results)
                print('GL Ref found in ' + cmf_list[a] + ' sheets ' + sheets[d])
              else:
                #audit_log('Finance SOA - GL No not found in ' + cmf_list[b] + ' sheets ' + sheets[d], soarpa_base)
                print('GL Ref not found in ' + cmf_list[a] + ' sheets ' + sheets[d])
                pass
            else:
              continue
        else:
          print("Updating SOA document: %s cannot be done. File is currently still LOCKED / USED." % filelocation)
          audit_log("Updating SOA document: %s cannot be done. File is currently still LOCKED / USED." % filelocation, 'Begin Updating the Hospital SOA File', base)
          pass
        continue

    print('# Perform merging')
    if final_df.empty == False:
      print('Testing merge')
      soa_df = pd.merge(soa_df, final_df[['GL Ref','Insurance', 'Submission Date', 'Bank-in Date', 'Amount (RM)']], how='left', left_on=['GL Ref'], right_on=['GL Ref'])
      insurance_backup = soa_df['Insurance_y']
      submission_backup = soa_df['Submission Date_y']
      bankinDate_backup = soa_df['Bank-in Date_y']
      amount_backup = soa_df['Amount (RM)_y']
      soa_df = soa_df.drop(['Insurance_y', 'Submission Date_y', 'Bank-in Date_y', 'Amount (RM)_y'], axis=1)
      soa_df['Insurance_x'] = insurance_backup
      soa_df['Submission Date_x'] = submission_backup
      soa_df['Bank-in Date_x'] = bankinDate_backup
      soa_df['Amount (RM)_x'] = amount_backup
      soa_df.rename(columns={'Insurance_x': 'Insurance', 'Submission Date_x': 'Submission Date', 'Bank-in Date_x': 'Bank-in Date', 'Amount (RM)_x':'Amount (RM)'}, inplace = True)
      print('Complete merge')
    else:
      #audit_log('Finance SOA - GL No does not exist in special client master file', soarpa_base)
      print('GL Ref does not exist')

    return soa_df

#!/usr/bin/python
# FINAL SCRIPT updated as of 12th June 2020
# Workflow - Finance Update CML
# Version 1

# Declare Python libraries needed for this script
import openpyxl
import re
import time
from datetime import datetime as dt
import pandas as pd
import numpy as np
from openpyxl import Workbook
from transformation.reqdca_excel_manipulation import *
from transformation.update_cmldb import *
#from transformation.update_cmlfile import *
from utils.audit_trail import audit_log
from utils.notification import send
from utils.logging import logging

##client AXA affin, MSIG, Allianz general

def preprosessing_cmlinfo(myclients, marcfile, paymenlisting, cmlrpa_base):
  try:

    # Database connection
    print("- Connecting to database")
    conn_mysql = MySqlConnector()
    cur_mysql = conn_mysql.cursor()

    # Load user files as workbooks
    wbmarc = load_workbook(marcfile, read_only = True)
    sheets_marc = wbmarc.sheetnames

    wbpaylist = load_workbook(paymenlisting, read_only = True)
    sheets_plist = wbpaylist.sheetnames

    i = 0
    for i in range(len(myclients)):

      # Calling Stored Procedure'
      parameters = [str(myclients['Client Name'][i]),]
      stored_proc = cur_mysql.callproc('dca_query_marc_for_client_id', parameters)
      for z in cur_mysql.stored_results():
        results = z.fetchall()
      fetched_result = pd.DataFrame(results)
      
      if fetched_result.empty != True and len(fetched_result) == 1:
        fetched_result.columns = ['inptCaseMmInsurerId', 'inptCaseMmInsurerParentId', 'inptCaseMmInsurerName']
        client_id = int(fetched_result['inptCaseMmInsurerId'][0])
      elif fetched_result.empty != True and len(fetched_result) >= 1:
        audit_log("Finance Update CML - Unable to get Client ID. Too many results. Ensure client name is specific.", "Completed...", cml_base)

      # Set the file location
      filelocation = myclients['Filenames'][i]
      print(filelocation)

      # Provide the name of the range of the header to be extracted
      first_header = 'Adm. No'
      second_header = 'Bank-in Date'
      last_header = '''Date of OR Rec'd'''
      get_header = 'Sage ID'

      wb = load_workbook(filelocation, read_only = False, keep_vba = True, data_only = False, keep_links = True)
      sheets = wb.sheetnames

      # Iterate the sheets available in the workbook
      a = 0
      for a in range(len(sheets)):

        print(sheets[a])

        # Get the cell position of the first and last header
        start_cell = search_excel(first_header, a, wb)
        end_cell = search_excel(last_header, a, wb)
        get_cell = search_excel(get_header, a, wb)

        if get_cell != [1, 5]:
          get_cell = search_excel(get_header, a, wb)
          if get_cell != []:
            pass
          else:
            get_cell = search_excel('SAGE DOC ID', a, wb)
        else:
          pass

        # Read Workbook as dataframe
        #searched_df = pd.read_excel(filelocation, sheet_name = a, skiprows = start_cell[1]-1, usecols = get_cell[0] + ':' + end_cell[0])
        searched_df = pd.read_excel(filelocation, sheet_name = a, skiprows = start_cell[1]-1)

        # Get the index to rename the column name in dataframe
        search_ls = list(searched_df.columns)
        find_column = ['Cheque Details', 'Hospital Bills Details', 'AAN Settlement Details', 'TPA Billing Details', 'Reimbursement Details', 'Adjustment Details']
        index_ls = []
        c = 0
        for c in range(len(find_column)):
          if find_column[c] in search_ls:
            column_index = []
            column_index.append(search_ls.index(find_column[c]))
            column_index.append(search_ls.index(find_column[c]) + 1)
            column_index.append(search_ls.index(find_column[c]) + 2)
            index_ls.append(column_index)
          else:
            pass
        
        if 'Hospital OR Details' in search_ls:
          column_index = []
          column_index.append(search_ls.index('Hospital OR Details'))
          column_index.append(search_ls.index('Hospital OR Details') + 1)
          column_index.append(search_ls.index('Hospital OR Details') + 2)
          if '''Date of OR Rec'd''' in search_ls:
            pass
          else:
            column_index.append(search_ls.index('Hospital OR Details') + 3)
            column_index.append(search_ls.index('Hospital OR Details') + 4)
          index_ls.append(column_index)
        else:
          pass
        
        if 'Designated  / Borrowing Bank Account' in search_ls:
          column_index = []
          column_index.append(search_ls.index('Designated  / Borrowing Bank Account'))
          column_index.append(search_ls.index('Designated  / Borrowing Bank Account') + 1)
          column_index.append(search_ls.index('Designated  / Borrowing Bank Account') + 2)
          column_index.append(search_ls.index('Designated  / Borrowing Bank Account') + 3)
          column_index.append(search_ls.index('Designated  / Borrowing Bank Account') + 4)
          column_index.append(search_ls.index('Designated  / Borrowing Bank Account') + 5)
          column_index.append(search_ls.index('Designated  / Borrowing Bank Account') + 6)
          index_ls.append(column_index)
        else:
          pass
        
        if 'Tax Invoice Details' in search_ls:
          column_index = []
          column_index.append(search_ls.index('Tax Invoice Details'))
          column_index.append(search_ls.index('Tax Invoice Details') + 1)
          column_index.append(search_ls.index('Tax Invoice Details') + 2)
          column_index.append(search_ls.index('Tax Invoice Details') + 3)
          column_index.append(search_ls.index('Tax Invoice Details') + 4)
          index_ls.append(column_index)
        else:
          pass
        
        # Rename the column name in dataframe
        correct_column_ls = [['Cheque Details - Bank-in Date', 'Cheque Details - Chq.No.', 'Cheque Details - Amount(RM)']
                            , ['Hospital Bills Details - Date', 'Hospital Bills Details - Bill No.', 'Hospital Bills Details - Amount(RM)']
                            , ['AAN Settlement Details - Date', 'AAN Settlement Details - Chq.No.', 'AAN Settlement Details - Amount(RM)']
                            , ['TPA Billing Details - Date', 'TPA Billing Details - Inv. No.', 'TPA Billing Details - Amount(RM)']
                            , ['Reimbursement Details - Date', 'Reimbursement Details - Chq.No.', 'Reimbursement Details - Amount(RM)']
                            , ['Adjustment Details - Date', 'Adjustment Details - Doc.No', 'Adjustment Details - Amount(RM)']
                            , ['Hospital OR Details - OR Date', 'Hospital OR Details - OR No.', 'Hospital OR Details - Amount(RM)', 'Hospital OR Details - Date of OR Rec\'d', 'Hospital OR Details - Subm.Date']
                            , ['Designated / Borrowing Bank Account - Bank 1', 'Designated / Borrowing Bank Account - Bank 2', 'Designated / Borrowing Bank Account - Bank 3', 'Designated / Borrowing Bank Account - Bank 4', 'Designated / Borrowing Bank Account - Bank 5', 'Designated / Borrowing Bank Account - Bank 6', 'Designated / Borrowing Bank Account - Bank 7']
                            , ['Tax Invoice Details - Date', 'Tax Invoice Details - Inv No', 'Tax Invoice Details - Case Fee', 'Tax Invoice Details - GST Amount', 'Tax Invoice Details - Total']]

        d = 0
        for d in range(len(index_ls)):
          e = 0
          for e in range(len(index_ls[d])):
            if searched_df.columns[index_ls[d][0]] == 'Cheque Details' or searched_df.columns[index_ls[d][0]] == 'Cheque Details - Bank-in Date':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            elif searched_df.columns[index_ls[d][0]] == 'Hospital Bills Details' or searched_df.columns[index_ls[d][0]] == 'Hospital Bills Details - Date':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            elif searched_df.columns[index_ls[d][0]] == 'AAN Settlement Details' or searched_df.columns[index_ls[d][0]] == 'AAN Settlement Details - Date':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            elif searched_df.columns[index_ls[d][0]] == 'TPA Billing Details' or searched_df.columns[index_ls[d][0]] == 'TPA Billing Details - Date':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            elif searched_df.columns[index_ls[d][0]] == 'Reimbursement Details' or searched_df.columns[index_ls[d][0]] == 'Reimbursement Details - Date':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            elif searched_df.columns[index_ls[d][0]] == 'Adjustment Details' or searched_df.columns[index_ls[d][0]] == 'Adjustment Details - Date':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            elif searched_df.columns[index_ls[d][0]] == 'Hospital OR Details' or searched_df.columns[index_ls[d][0]] == 'Hospital OR Details - OR Date':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            elif searched_df.columns[index_ls[d][0]] == 'Designated  / Borrowing Bank Account' or searched_df.columns[index_ls[d][0]] == str(searched_df[searched_df.columns[index_ls[d][0]]].iloc[0]) or searched_df.columns[index_ls[d][0]] == 'Designated / Borrowing Bank Account - Bank 1':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            elif searched_df.columns[index_ls[d][0]] == 'Tax Invoice Details' or searched_df.columns[index_ls[d][0]] == 'Tax Invoice Details - Date':
              searched_df = searched_df.rename(columns={searched_df.columns[index_ls[d][e]] : correct_column_ls[d][e]})
            else:
              pass
        
        col_a = int(len(searched_df[pd.notnull(searched_df['Adm. No'])]))
        col_b = int(len(searched_df[pd.notnull(searched_df['Hospital Bills Details - Date'])]))
        
        if col_a != col_b:
          searched_df = pd.DataFrame(searched_df[pd.isnull(searched_df['Hospital Bills Details - Date'])])
          searched_df = searched_df[pd.notnull(searched_df['Adm. No'])]
          
          # Filter date ranges
          start_date = str(datetime.datetime.strptime(jsondf['Start'][0], "%Y-%m-%d").strftime("%Y-%m-%d"))
          end_date = str(datetime.datetime.strptime(jsondf['End'][0], "%Y-%m-%d").strftime("%Y-%m-%d"))
          searched_df = searched_df.loc[(searched_df['OB Registered Date'] > start_date) & (searched_df['OB Registered Date'] <= end_date)]
          
          # VLOOKUP
          downloadmarc_df = pd.read_excel(marcfile, sheet_name = 0)
          merged_df = pd.merge(left = downloadmarc_df, right = searched_df, left_on = downloadmarc_df['PONBR'], right_on = searched_df['Adm. No'].astype(int))
          merged_df = merged_df.drop(columns = ['key_0'])
          
          if merged_df.empty == True:
            print("Finance Update CML - RPA task execution completion notification. No matched data between CML file [ %s ] and downloaded report from MARC." % (filelocation))
            audit_log("Finance Update CML - RPA task execution completion notification. No matched data between CML file [ %s ] and downloaded report from MARC." % filelocation, "Completed...", cmlrpa_base)
          else:
            merged_df['Hospital Bills Details - Amount(RM)'] = merged_df['AMTGROSTOT']
            merged_df['Hospital Bills Details - Bill No.'] = merged_df['IDINVC']
            merged_df['Hospital Bills Details - Date'] = merged_df['VALUE.1'].astype(int)
            
            # VLOOKUP with Payment Listing
            paymenlisting_df = pd.read_excel(paymenlisting, sheet_name = 0)
            merged_df = pd.merge(left = paymenlisting_df, right = merged_df, left_on = paymenlisting_df['GLREF'], right_on = merged_df['Adm. No'].astype(object))
            if merged_df.empty == True:
              print("Finance Update CML - RPA task execution completion notification. No matched data between CML file [ %s ] and Payment Listing [ %s ]." % (filelocation, paymenlisting))
              audit_log("Finance Update CML - RPA task execution completion notification. No matched data between CML file [ %s ] and Payment Listing [ %s ]." % (filelocation, paymenlisting), "Completed...", cml_base)
            else:
              b = 0
              for b in range(len(merged_df['IDRMIT'])):
                merged_df.loc[b, 'IDRMIT'] = str(merged_df['IDRMIT'][b]).zfill(12)
            
            k = 0
            for k in range(len(merged_df['BANKCODE'])):
              if merged_df['BANKCODE'][k] == 'DBMY21':
                mybankname1 = str(merged_df['BANKCODE'][k])
                merged_df.rename(columns = {'Designated / Borrowing Bank Account - Bank 1':mybankname1}, inplace = True)
                merged_df.loc[k, 'DBMY21'] = merged_df['AMTPAYM'][k]
              elif merged_df['BANKCODE'][k] == 'DBMY39':
                merged_df.loc[k, 'DBMY39'] = merged_df['AMTPAYM']
              elif merged_df['BANKCODE'][k] == merged_df['BANKCODE'][k]:
                mybankname1 = str(merged_df['BANKCODE'][k])
                merged_df.rename(columns = {'Bank 1':mybankname1}, inplace = True)
                merged_df.loc[k, str(mybankname1)] = merged_df['AMTPAYM'][k]
              elif merged_df['BANKCODE'][k] == merged_df['BANKCODE'][k] != mybankname1:
                mybankname2 = str(merged_df['BANKCODE'][k])
                merged_df.rename(columns = {'Bank 2':mybankname2}, inplace = True)
                merged_df.loc[k, str(mybankname2)] = merged_df['AMTPAYM'][k]
              elif merged_df['BANKCODE'][k] == merged_df['BANKCODE'][k] != mybankname2:
                mybankname3 = str(merged_df['BANKCODE'][k])
                merged_df.rename(columns = {'Bank 3':mybankname3}, inplace = True)
                merged_df.loc[k, str(mybankname3)] = merged_df['AMTPAYM'][k]
              elif merged_df['BANKCODE'][k] == merged_df['BANKCODE'][k] != mybankname3:
                mybankname4 = str(merged_df['BANKCODE'][k])
                merged_df.rename(columns = {'Bank 4':mybankname4}, inplace = True)
                merged_df.loc[k, str(mybankname4)] = merged_df['AMTPAYM'][k]
              else:
                merged_df.loc[k, 'Bank 5'] = merged_df['AMTPAYM'][k]

            # Clean up dataframe for final use
            merged_df = merged_df.drop(columns = ['key_0', 'BATCHTYPE', 'CNTBTCH_x', 'CNTRMIT', 'CNTLINE', 'IDINVC_x', 'GLREF', 'DATERMIT', 'IDRMIT', 'AMTPAYM', 'DATEINVC_x', 'BANKCODE', 'Co./Last Name',
                                                  'CNTBTCH_y', 'CNTITEM', 'IDVEND', 'PONBR', 'IDINVC_y', 'TEXTTRX', 'ORDRNBR', 'INVCDESC', 'DATEINVC_y', 'CODECURN', 'EXCHRATEHC', 'TAXCLASS1', 'AMTGROSTOT',
                                                  'IDGLACCT', 'OPTFIELD', 'VALUE', 'OPTFIELD.1', 'VALUE.1', 'Unnamed: 19'])

            merged_df = merged_df.rename(columns={' Under / (Over) Claimed                 fr Client                  (AA-J-AM+AP)' : 'Under / (Over) Claimed fr Client (AA-J-AM+AP)'})
            merged_df = merged_df.rename(columns={'Under / (Over) Reimbursed by Client                   (AB-Y-AN+AQ)' : 'Under / (Over) Reimbursed by Client (AB-Y-AN+AQ)'})
            merged_df = merged_df.rename(columns={'Under / (Over) Paid to Hosp      (AB - AG)' : 'Under / (Over) Paid to Hosp (AB - AG)'})

            check_col_ls = list(merged_df.columns)
            if 'Tax Invoice Details - Date' in check_col_ls:
              pass
            else:
              merged_df['Tax Invoice Details - Date'] = np.nan
              merged_df['Tax Invoice Details - Inv No'] = np.nan
              merged_df['Tax Invoice Details - Case Fee'] = np.nan
              merged_df['Tax Invoice Details - GST Amount'] = np.nan
              merged_df['Tax Invoice Details - Total'] = np.nan

            merged_df = merged_df.replace(np.nan, '', regex=True)


            # Update CML file
            #update_cml(merged_df, sheets[a], filelocation, cmlrpa_base)

            # Update CML file in the database
            update_cmldb(merged_df, sheets[a], client_id, cmlrpa_base)

        else:
          print("Finance CML Update - Can't process [ %s ] sheet for file [ %s ]" % (sheets[a], filelocation))
          continue

    wbmarc.close
    wbpaylist.close

  except Exception as error:
    pass

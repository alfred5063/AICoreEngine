#!/usr/bin/python
# FINAL SCRIPT updated as of 8th April 2020
# Workflow - Finance-AR
# Version 2

# Declare Python libraries needed for this script
import pandas as pd
from openpyxl import load_workbook
import sys
from datetime import datetime
import os
from utils.Session import session
from utils.audit_trail import audit_log
from utils.logging import logging
from transformation.reqdca_excel_manipulation import *
from automagica import *


def mapping_sage(template_type, template_excel, jasper_excel, now, base):
  try:
    audit_log("Mapping Process", "Mapping Information in Jasper File to Sage File", base)
    print("Mapping Information in Jasper File to Sage File")
    year = now.year
    month = now.month
    day = now.day

    jasper_df = pd.read_excel(jasper_excel, skiprows = 5, usecols = 'A:AQ')
    template_df = sage_df_list = invoice_df = invoice_details_df = invoice_optional_df = invoice_details_optional_df = invoice_payment_schedules = pd.DataFrame()
    wb = load_workbook(template_excel)

    #Assign a dataframe for each sheet
    for sheet in wb.worksheets:
      template_df = pd.DataFrame(pd.read_excel(template_excel, sheet_name = sheet.title, nrows = 0, usecols = list(range(sheet.max_column+1))))
      sage_df_list.append(template_df)
      if sheet.title.lower() == 'invoices':
        invoice_df = template_df
      if sheet.title.lower() == 'invoice_details':
        invoice_details_df = template_df
      if sheet.title.lower() == 'invoice_optional_fields' :
        invoice_optional_df = template_df
      if sheet.title.lower() == 'invoice_detail_optional_fields':
        invoice_details_optional_df = template_df
      if sheet.title.lower() == 'invoice_payment_schedules':
        invoice_payment_schedules = template_df

    jasper_df['Cost/Invoiced Amt'].fillna(0, inplace=True)

    #Divide the main dataframe into several dataframe grouped by the Tix ID 
    grouped = list(jasper_df.groupby(['Tix ID'], sort=False))

    try:

      #The process if template is MY Sage Report
      if template_type == 'sage':
        invoice_optional_field_list = ['00TYPESERVIC','05INSURED','09POLICYNO','10DATEOCCUR','11DATEADMIN','12DATEDISCHR','12DATEOASST','12DATEOCC','13FREETEXT1','14FREETEXT2','15FREETEXT3']
        invoice_optional_value_list = [None,'ORDRNBR','Client Ref Num','Tix Created Date','Tix Created Date','Ops Completed Date',None,'Ops Completed Date',None,None,None]


        for i in range(len(grouped)):
  
          #temporary dataframe to reset the index of the dataframe
          df = pd.DataFrame(grouped[i][1])
          df = df.reset_index().drop(columns='index')

          #Total invoice gathered from cost/invoiced amount of same Tix ID
          total_invoice_amount = 0.00
  
          for j in range(df.shape[0]):
            total_invoice_amount = total_invoice_amount + df.iloc[j]['Cost/Invoiced Amt']

            #Set idaccrev (default is '4710000-IMA' while Client AXA is based on their Tix ID)
            if('axa' in df.iloc[j]['Client Name'].lower()):
              idacctrev = None
            else:
              idacctrev = '4710000-IMA'

            if(pd.isna(df.iloc[j]['Provider Name'])):
              comment = "%s, %s" % (df.iloc[j]['Cost Type'], df.iloc[j]['Invoice Num'])
            else:
              comment = "%s, %s,%s" % (df.iloc[j]['Cost Type'], df.iloc[j]['Invoice Num'], df.iloc[j]['Provider Name'])

            #Append new row in Invoice Details dataframe
            invoice_details_df=invoice_details_df.append({'CNTBTCH': 1193, #value unknown
                                                        'CNTITEM': i+1,
                                                        'CNTLINE':(j+1)*20,
                                                        'TEXTDESC':df.iloc[j]['Invoice Num'],
                                                        'AMTEXTN': df.iloc[j]['Cost/Invoiced Amt'],
                                                        'AMTTXBL': df.iloc[j]['Cost/Invoiced Amt'],
                                                        'BASETAX1': df.iloc[j]['Cost/Invoiced Amt'],
                                                        'TAXSTTS1': 1,
                                                        'IDACCTREV': idacctrev,
                                                        'COMMENT':comment},
                                                        ignore_index=True)

            #Append new row in Details Optional dataframe
            invoice_details_optional_df=invoice_details_optional_df.append({'CNTBTCH': 1193, #value unknown
                                                                            'CNTITEM': i+1,
                                                                            'CNTLINE':(j+1)*20,
                                                                            'OPTFIELD':'CASENO',
                                                                            'VALUE': grouped[i][1].iloc[0]['Tix ID']},
                                                                            ignore_index=True)

            #Append new row in Details Optional dataframe for TP
            invoice_details_optional_df=invoice_details_optional_df.append({'CNTBTCH': 1193, #value unknown
                                                                            'CNTITEM': i+1,
                                                                            'CNTLINE':(j+1)*40,
                                                                            'OPTFIELD':'COSTINGID',
                                                                            'VALUE': grouped[i][1].iloc[j]['Costing Id']},
                                                                            ignore_index=True)

          #Store total invoice amount in grouped tuple respectively
          l=list(grouped[i])
          l.append(total_invoice_amount)
          grouped[i]=tuple(l)

          #Append new row in invoice dataframe
          invoice_df=invoice_df.append({'CNTITEM': i+1,
                                        'ORDRNBR': grouped[i][1].iloc[0]['Member First Name']+" "+grouped[i][1].iloc[0]['Member Last Name'],
                                        'INVCDESC': grouped[i][1].iloc[0]['Invoice Num'],
                                        'CUSTPO': grouped[i][1].iloc[0]['Tix ID'],
                                        'TEXTTRX': 1,
                                        'IDCUST':grouped[i][1].iloc[0]['Client Name'],
                                        'DATEINVC':"%s/%s/%s" %(day,month,year),
                                        'DATEASOF':"%s/%s/%s" %(day,month,year),
                                        'FISCYR':year,
                                        'FISCPER':month,
                                        'CODETAXGRP':'SSTNA',
                                        'CODETAX1':'SSTNA',
                                        'TAXSTTS1':1,
                                        'AMTTXBL':grouped[i][2],
                                        'DATEDUE':"%s/%s/%s" %(day,(month%12)+1,year),
                                        'DATERATE':"%s/%s/%s" %(day,month,year),
                                        'AMTINVCTOT':grouped[i][2],
                                        'CODECURNRC':'MYR',
                                        'RATEDATERC':"%s/%s/%s" %(day,month,year),
                                        'DATEBUS':"%s/%s/%s" %(day,month,year)},
                                        ignore_index=True)

         #Append new row into Invoice Optional Dataframe
          for k in range(len(invoice_optional_field_list)):
            value = ''
            if(k==1):
              value = invoice_df.iloc[i][invoice_optional_value_list[k]]
            elif(invoice_optional_value_list[k]==None):
              value = None
            elif('timestamps.Timestamp' in str(type(grouped[i][1].iloc[0][invoice_optional_value_list[k]]))):
              value = grouped[i][1].iloc[0][invoice_optional_value_list[k]].strftime('%Y%m%d')
            else:
              value = grouped[i][1].iloc[0][invoice_optional_value_list[k]]
            invoice_optional_df=invoice_optional_df.append({'CNTITEM': i+1,
                                                            'OPTFIELD': invoice_optional_field_list[k],
                                                            'VALUE': value},
                                                            ignore_index=True)

      #Process for Sgpore Template
      elif template_type == 'sgpore':
        invoice_optional_field_list = ['00DISCLAIMER','01CONDESC1','01CONDESC2','01CONDESC3','04VEHNO','05INSURED','06REFERENCE','07LOCATION','08CLAIM','09POLICYNO','10DATEOCCUR','11DATEADMIN','12DATEDISCHR']
        invoice_optional_value_list = ['REMARK01',None,None,None,None,'ORDRNBR',None,None,None,'Client Ref Num',None,'Tix Created Date','Ops Completed Date']
        invoice_details_optional_field_list = ['CASENO','CONTRACTNO','DEPT','TP','ZPROVCODE','ZPROVINVNO']
        invoice_details_optional_value_list = ['Tix ID',4900193,None,None,None,None]

      
        for i in range(len(grouped)):
  
          #temporary dataframe to reset the index of the dataframe
          df = pd.DataFrame(grouped[i][1])
          df = df.reset_index().drop(columns='index')

          #Total invoice gathered from cost/invoiced amount of same Tix ID
          total_invoice_amount = 0.00
  
          for j in range(df.shape[0]):
            total_invoice_amount = total_invoice_amount + df.iloc[j]['Cost/Invoiced Amt']

            if(pd.isna(df.iloc[j]['Provider Name'])):
              comment = "%s" % df.iloc[j]['Invoice Num']
            else:
              comment = "%s, %s" % (df.iloc[j]['Invoice Num'], df.iloc[j]['Provider Name'])

            #Append new row in Invoice Details dataframe
            invoice_details_df=invoice_details_df.append({'CNTBTCH': 1979, #value unknown
                                                          'CNTITEM': i+1,
                                                          'CNTLINE':(j+1)*20,
                                                          'TEXTDESC':"%s, %s" %(df.iloc[j]['Main Service'], df.iloc[j]['Cost Type']),
                                                          'AMTEXTN': df.iloc[j]['Cost/Invoiced Amt'],
                                                          'AMTTXBL': df.iloc[j]['Cost/Invoiced Amt'],
                                                          'BASETAX1': df.iloc[j]['Cost/Invoiced Amt'],
                                                          'TAXSTTS1': 2,
                                                          'IDACCTREV': 4110600,
                                                          'COMMENT':comment},
                                                          ignore_index=True)

          #Store total invoice amount in grouped tuple respectively
          l=list(grouped[i])
          l.append(total_invoice_amount)
          grouped[i]=tuple(l)

          #Append new row in invoice dataframe
          invoice_df=invoice_df.append({'CNTBTCH': 1979,
                                        'CNTITEM': i+1,
                                        'IDCUST': grouped[i][1].iloc[0]['Client Name'],
                                        'TEXTTRX': 1,
                                        'ORDRNBR': grouped[i][1].iloc[0]['Member First Name']+" "+grouped[i][1].iloc[0]['Member Last Name'],
                                        'CUSTPO': grouped[i][1].iloc[0]['Tix ID'],
                                        'INVCDESC': grouped[i][1].iloc[0]['Invoice Num'],
                                        'DATEINVC':"%s/%s/%s" %(day,month,year),
                                        'DATEASOF':"%s/%s/%s" %(day,month,year),
                                        'FISCYR':year,
                                        'FISCPER':month,
                                        'CODECURN': 'SGD',
                                        'EXCHRATEHC': 1,
                                        'TERMCODE': 30,
                                        'DATEDUE':"%s/%s/%s" %(day,(month%12)+1,year),
                                        'CODETAXGRP':'TAXSGD',
                                        'CODETAX1':'TAXSGD',
                                        'TAXSTTS1':2,
                                        'AMTTXBL':grouped[i][2],
                                        'AMTNOTTXBL': 0,
                                        'AMTINVCTOT': grouped[i][2],
                                        'DATERATE':"%s/%s/%s" %(day,month,year),
                                        'CODECURNRC':'MYR',
                                        'RATERC': 1,
                                        'RATEDATERC':"%s/%s/%s" %(day,month,year),
                                        'DATEBUS':"%s/%s/%s" %(day,month,year)},
                                        ignore_index=True)

          #Append new row into Invoice Optional Dataframe
          for k in range(len(invoice_optional_field_list)):
            value = ''
            if(k==5):
              value = invoice_df.iloc[i][invoice_optional_value_list[k]]
            elif(invoice_optional_value_list[k]==None):
              value = None
            elif('REMARK01' in invoice_optional_value_list[k]):
              value = invoice_optional_value_list[k]
            else:
              value = grouped[i][1].iloc[0][invoice_optional_value_list[k]]
            invoice_optional_df=invoice_optional_df.append({'CNTBTCH': 2003,
                                                            'CNTITEM': i+1,
                                                            'OPTFIELD': invoice_optional_field_list[k],
                                                            'VALUE': value},
                                                            ignore_index=True)

          #Append new row into Invoice Details Optional Dataframe
          for l in range(len(invoice_details_optional_field_list)):
            value = ''
            if(invoice_details_optional_value_list[l]==None):
              value = None
            elif('int' in str(type(invoice_details_optional_value_list[l]))):
              value = invoice_details_optional_value_list[l]
            else:
              value = grouped[i][1].iloc[0][invoice_details_optional_value_list[l]]
            invoice_details_optional_df=invoice_details_optional_df.append({'CNTBTCH': '2003',
                                                                            'CNTITEM': i+1,
                                                                            'CNTLINE': 20,
                                                                            'OPTFIELD': invoice_details_optional_field_list[l],
                                                                            'VALUE': value},
                                                                            ignore_index=True)

      audit_log("Mapping Process", "Completed... All information is mapped", base)

    except Exception as error:
      logging("Error in Mapping Process", error, base)
      print("Error in Mapping Process " + error)

    audit_log("Update Sage", "Update Sage Report based on the mapping", base)

    print("Update Sage Report based on the mapping")

    sage_df_list = [invoice_df,invoice_details_df,invoice_payment_schedules,invoice_optional_df,invoice_details_optional_df]
    
    for i in range(len(wb.worksheets)):
      try:
        if(len(sage_df_list) != len(wb.worksheets) and sage_df_list[i].empty):
          sage_df_list.pop(i)
        start_cell = 2
        for j in range(sage_df_list[i].shape[0]):     
          update_entry(wb.worksheets[i], sage_df_list[i], start_cell, j, 1)
          start_cell = start_cell + 1
      except Exception as error:
        print(error)
        continue

    save_and_close(wb, template_excel)

  except Exception as error:
    logging("Error in Updating Sage Report", error, base)
    print("Error in Updating Sage Report ",error)



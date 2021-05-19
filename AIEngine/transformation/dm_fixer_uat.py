#!/usr/bin/python
# FINAL SCRIPT updated as of 8th Jan 2021
# Workflow - STP-DM - UAT

# Declare Python libraries needed for this script
import pandas as pd
import numpy as np
import xlwt
import shutil
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from transformation.reqdca_excel_manipulation import *
from transformation.dm_fixer import *
from directory.files_listing_operation import getListOfFiles
from directory.get_filename_from_path import *
from utils.audit_trail import audit_log
from utils.logging import logging
from transformation.dm_file_preparation import *
import win32com.client as win32
import pythoncom

def read_csv_skiprows(path):
  with open(path) as f:
      lines = f.readlines()
      print(lines)
      #get list of all possible lins starting by No
      num = [i for i, l in enumerate(lines) if l.startswith("No")]
      #if not found value return 0 else get first value of list subtracted by 1
      num = 0 if len(num) == 0 else num[0] - 1
      return num

def clean_file(onefile):
  df = pd.DataFrame(onefile.replace(np.nan, '', regex = True))
  df2 = pd.DataFrame(df.replace('nan', '', regex = False))
  df3 = pd.DataFrame(df2.replace('NaT', '', regex = False))
  if 'No' in df3.columns:
    df3.dropna(subset=['No'], inplace=True)
    df3['No'] = df3['No'].astype(int)
  elif 'NO' in df3.columns:
    df3.dropna(subset=['NO'], inplace=True)
    df3['NO'] = df3['NO'].astype(int)
  return df3


def fixering(df, wb, sheets, fixer, sheet_num, keywerd):
  i = 0
  start_cell = search_excel(str(keywerd), sheet_num, wb)
  if start_cell != None:
    last_row = int(search_last_row(start_cell, fixer, sheet_num, wb, str(keywerd)) - 1)
  else:
    start_cell = search_excel(str(keywerd), sheet_num, wb)
    last_row = int(search_last_row(start_cell, fixer, sheet_num, wb, str(keywerd)) - 1)
  for i in range(len(df)):
    update_entry(wb.active, df.loc[df.index == i], last_row + i, i, start_column = 1)
  wb.save(str(fixer))
  wb.close()


#def openfile(fixer):
#  fixer = str(fixer).replace(str('\\'), str('\\\\'))
#  print(fixer)
#  excel = win32.gencache.EnsureDispatch('Excel.Application')
#  try:
#    print(str(fixer))
#    print("A")
#    wb = excel.Workbooks.Open(str(fixer))
#    print(wb)
#  except Exception as error:
#    print(str(fixer))
#    print("B")
#    print("cannot launch workbook")
#    print(error)
#    print("Error in opening in-memory FIXER file.")
#  excel.EnableEvents = False
#  excel.DisplayAlerts = False
#  excel.Visible = False
#  wb.Save()
#  wb.Close(True)
#  excel.Quit()


def revising(fixer, sheetname, client):
  wb = load_workbook(fixer, data_only=True)
  wb.active = wb.worksheets.index(wb[sheetname])
  dict = dir_string_fixer(client,sheetname)
  searched_df = pd.read_excel(fixer, sheet_name = int(wb.worksheets.index(wb[sheetname])), dtype = dict)
  wb.close()
  return searched_df

#source = str(ARCHIVED + "\\" + tail_input)
#filename = FILENAME
#result = RESULTS
#failed = FAILED
#client = extracted_clist['Client'][l]
#fixer = str(TEMP + "\\" + tail_fixer)
#prefix = PREFIX
#stpdm_base

def skiprows_input(source):
  wb = load_workbook(source)
  sheets = wb.sheetnames
  firstrow = ''
  for v in range(len(sheets)):
    wb.active = wb.worksheets.index(wb[sheets[v]])
    sheet_num = int(wb.worksheets.index(wb[sheets[v]]))
    kwd = ['No','NO','No.','NO.','POLICY APPROVAL DATE']
    for i in range(len(kwd)):
      try:
        firstrow = int(search_excel(kwd[i], sheet_num, wb)[1]) - int(1)
        print('firstrow')
        print(firstrow)
      except:
        wb.close()
        pass
      continue
  wb.close()
  return firstrow

def read_file(source, filename, result, failed, client, fixer, prefix, stpdm_base):
  wb = load_workbook(fixer)
  sheets = wb.sheetnames
  firstrow = ''
  keywerd = ''
  for v in range(len(sheets)):
    if 'raw' in str(sheets[v].lower()):
      wb.active = wb.worksheets.index(wb[sheets[v]])
      sheet_num = int(wb.worksheets.index(wb[sheets[v]]))
      kwd = ['No','NO','No.','NO.']
      for i in range(len(kwd)):
        try:
          firstrow = int(search_excel(kwd[i], sheet_num, wb)[1]) - int(1)
          keywerd = kwd[i]
        except:
          pass
        continue
    #firstrow_input = skiprows_input(source)
    #print('firstrow')
    #print(firstrow_input)
    try:
      if '.xlsx' in str(source.lower()):
        firstrow_input = skiprows_input(source)
        dict = dir_string(client)
        onefile = pd.read_excel(source, dtype = dict, skiprows = firstrow_input).astype(str)
        col_tag = True
      elif '.xls' in str(source.lower()):
        dict = dir_string(client)
        onefile = pd.read_excel(source, dtype = dict).astype(str)
        thiscol = onefile.columns[0]
        onefile = pd.read_excel(source, dtype = dict, skiprows = int(min(onefile[thiscol])) - 1).astype(str)
        col_tag = True
      elif '.csv' in str(source.lower()):
        try:
          #dict = dir_string(client)
          #onefile = pd.read_csv(source, encoding = "ISO-8859-1", skiprows = firstrow_input, dtype = dict).astype(str)
          onefile = pd.read_csv(source, encoding = "ISO-8859-1", skiprows = firstrow, dtype = dict).astype(str)
          col_tag = True
        except:
          #dict = dir_string(client)
          #onefile = pd.read_csv(source, encoding = "ISO-8859-1", skiprows = firstrow_input, dtype = str).astype(str)
          onefile = pd.read_csv(source, encoding = "ISO-8859-1", skiprows = firstrow, dtype = dict).astype(str)
          col_tag = True
      elif '.txt' in str(source.lower()):
        try:
          onefile = pd.read_csv(source, sep = '|', skiprows = 1, header = None)
          onefile = onefile.rename(columns = {0:'No'})
          col_tag = True
        except:
          onefile = pd.read_csv(source, sep = '|', header = None, skiprows = 1, error_bad_lines = False)
          onefile = onefile.rename(columns = {0:'No'})
          col_tag = True
        pass
    except Exception as error:
      col_tag = False
      shutil.move(source, str(failed), stpdm_base)
      print(error)
      print("Error in parsing columns")
      logging("STP-DM - No columns to parse from file.", error, stpdm_base)
      pass
    continue

  if onefile.empty != True and col_tag == True:
    head, tail = get_file_name(source)
    audit_log("STP-DM - File [ %s ] is not empty and will be processed in Fixer." %tail, "Completed...", stpdm_base)

    try:
      df = clean_file(onefile)
      try:
        print('start fixering')
        fixering(df, wb, sheets, fixer, sheet_num, keywerd)
        print('done fixering')
      except Exception as error:
        print("Error in FIXERing")
        print(error)
        logging("STP-DM - Error in FIXERing.", error, stpdm_base)
        pass

      try:
        fixer2 = str(fixer).replace(str('\\'), str('\\\\'))
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fixer2)
        print("A")
        print(wb)
        excel.EnableEvents = False
        excel.DisplayAlerts = False
        excel.Visible = False
        wb.Save()
        wb.Close(True)
        excel.Quit()
      except Exception as error:
        print("B")
        print("cannot launch workbook")
        print(error)
        print("Error in opening in-memory FIXER file.")
        logging("STP-DM - Error in opening in-memory FIXER file.", error, stpdm_base)
        pass

      try:
        if str("_IO_") in prefix:
          prefix_ip = prefix.replace("_IO_", "_IP_")
          prefix_op = prefix.replace("_IO_", "_OP_")
          for m in range(len(sheets)):
            if 'MARC' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC', client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              df2.to_excel(result + "\\" + prefix + filename + ".xlsx", sheet_name = 'MARC', index = False)
              with pd.ExcelWriter(result + "\\" + prefix + filename + ".xlsx", engine = 'openpyxl', mode = 'a') as writer:
                for m in range(len(sheets)):
                  if 'CARD' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m]), client)
                    df.to_excel(writer, sheet_name = str(sheets[m]), index = False)
                  elif 'RAW' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m].upper()), client)
                    df.to_excel(writer, sheet_name = str(sheets[m].upper()), index = False)
                writer.save()
            elif 'MARC (OP)' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC (OP)',client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              df2.to_excel(result + "\\" + prefix_op + filename + ".xlsx", sheet_name = 'MARC (OP)', index = False)
              with pd.ExcelWriter(result + "\\" + prefix_op + filename + ".xlsx", engine = 'openpyxl', mode = 'a') as writer:
                for m in range(len(sheets)):
                  if 'CARD' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m]), client)
                    df.to_excel(writer, sheet_name = str(sheets[m]), index = False)
                  elif 'RAW' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m].upper()), client)
                    df.to_excel(writer, sheet_name = str(sheets[m].upper()), index = False)
                writer.save()
            elif 'MARC (IP)' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC (IP)', client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              df2.to_excel(result + "\\" + prefix_ip + filename + ".xlsx", sheet_name = 'MARC (IP)', index = False)
              with pd.ExcelWriter(result + "\\" + prefix_ip + filename + ".xlsx", engine = 'openpyxl', mode = 'a') as writer:
                for m in range(len(sheets)):
                  if 'CARD' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m]), client)
                    df.to_excel(writer, sheet_name = str(sheets[m]), index = False)
                  elif 'RAW' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m].upper()), client)
                    df.to_excel(writer, sheet_name = str(sheets[m].upper()), index = False)
                writer.save()
        else:
          for m in range(len(sheets)):
            if 'MARC' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC',client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              df2.to_excel(result + "\\" + prefix + filename + ".xlsx", sheet_name = 'MARC', index = False)
            elif 'MARC (OP)' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC (OP)',client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              df2.to_excel(result + "\\" + prefix + filename + ".xlsx", sheet_name = 'MARC (OP)', index = False)
            elif 'MARC (IP)' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC (IP)',client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              df2.to_excel(result + "\\" + prefix + filename + ".xlsx", sheet_name = 'MARC (IP)', index = False)
          with pd.ExcelWriter(result + "\\" + prefix + filename + ".xlsx", engine = 'openpyxl', mode = 'a') as writer:
            for m in range(len(sheets)):
              if 'CARD' in str(sheets[m].upper()):
                df = revising(fixer, str(sheets[m]),client)
                df.to_excel(writer, sheet_name = str(sheets[m]), index = False)
              elif 'RAW' in str(sheets[m].upper()):
                df = revising(fixer, str(sheets[m].upper()),client)
                df.to_excel(writer, sheet_name = str(sheets[m].upper()), index = False)
            writer.save()
      except Exception as error:
        print("Error in Revising dataframe from FIXER")
        print(error)
        logging("STP-DM - Error in Revising dataframe from FIXER.", error, stpdm_base)
        pass

    except Exception as error:
      shutil.move(source, failed, stpdm_base)
      logging("STP-DM - Error in file cleaning.", error, stpdm_base)
      pass

  elif onefile.empty == True:
    shutil.move(source, failed, stpdm_base)

def cleanup_df(df):
  df1 = pd.DataFrame(df.replace('-', '', regex = False))
  return df1

def date_validation(df):
  for i in df.columns:
    try:
      kwd = ['Date', 'DATE', 'DOB', 'date']
      for j in range(len(kwd)):
        try:
          if kwd[j] in i:
            #df[i] = pd.to_datetime(df[i])
            df[i] = df[i].dt.strftime('%d/%m/%Y')
        except:
          pass
        continue
    except:
      pass
    continue
  df1 = pd.DataFrame(df.replace(np.nan, '', regex = True))
  df2 = pd.DataFrame(df1.replace('nan', '', regex = False))
  df3 = pd.DataFrame(df2.replace('NaT', '', regex = False))
  return df3

def dir_string(client):

  if 'LON' in client:
    # lst of column names which needs to be string
    lst_str_cols = ['Address2', 'Address3', 'I.C.No', 'Hand Phone', 'House Phone', 'Work Phone']
    # use dictionary comprehension to make dict of dtypes
    dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'STMB GROUP' in client:
    # lst of column names which needs to be string
    lst_str_cols = ['IC No (New)']
    # use dictionary comprehension to make dict of dtypes
    dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'STMB_IND' in client: #IP
    # lst of column names which needs to be string
    lst_str_cols = ['NRIC ID', 'Postal Code']
    # use dictionary comprehension to make dict of dtypes
    dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'PACIFIC' in client:
    # lst of column names which needs to be string
    lst_str_cols = ['IC', 'ADDRESS03', 'ADDRESS02']
    # use dictionary comprehension to make dict of dtypes
    dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'AGIC' in client:
    # lst of column names which needs to be string
    lst_str_cols = ['Address 2', 'Address 3', 'NRIC', 'Principal NRIC']
    # use dictionary comprehension to make dict of dtypes
    dict_dtypes = {x : 'str'  for x in lst_str_cols}

  return dict_dtypes

def dir_string_fixer(client,sheetname):

  if 'LON' in client:
    if 'RAW' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address2', 'Address3', 'I.C.No']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
    elif 'MARC' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address 2', 'Address 3', 'NRIC','Principal NRIC']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
    elif 'CARD' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address 2', 'Address 3','I.C. No']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'STMB GROUP' in client:
    if 'RAW' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['IC No (New)']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
    elif 'MARC' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address 2', 'Address 3', 'NRIC','Principal NRIC']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
    elif 'CARD' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address 2', 'Address 3','I.C. No']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'STMB_IND' in client:
    if 'RAW' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['NRIC ID', 'Postal Code']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
    elif 'MARC' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address 2', 'Address 3', 'NRIC','Principal NRIC']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
    elif 'CARD' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address 2', 'Address 3','I.C. No']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'PACIFIC' in client:
    if 'RAW' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['IC', 'ADDRESS03', 'ADDRESS02']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
    elif 'MARC' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address 2', 'Address 3', 'NRIC','Principal NRIC']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
    elif 'CARD' in sheetname:
      # lst of column names which needs to be string
      lst_str_cols = ['Address 2', 'Address 3','I.C. No']
      # use dictionary comprehension to make dict of dtypes
      dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'AGIC' in client:
    # lst of column names which needs to be string
    lst_str_cols = ['Address 2', 'Address 3', 'NRIC', 'Principal NRIC']
    # use dictionary comprehension to make dict of dtypes
    dict_dtypes = {x : 'str'  for x in lst_str_cols}
  elif 'GELM' in client:
    # lst of column names which needs to be string
    lst_str_cols = ['Address 2', 'Address 3', 'NRIC', 'Principal NRIC']
    # use dictionary comprehension to make dict of dtypes
    dict_dtypes = {x : 'str'  for x in lst_str_cols}

  return dict_dtypes

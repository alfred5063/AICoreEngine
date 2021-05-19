#!/usr/bin/python
# FINAL SCRIPT updated as of 12th April 2021
# Workflow - STP-DM

# Declare Python libraries needed for this script
import pandas as pd
import numpy as np
import xlwt
import shutil
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from transformation.reqdca_excel_manipulation import *
from transformation.dm_fixer import *
from directory.files_listing_operation import getListOfFiles
from directory.get_filename_from_path import *
from utils.audit_trail import audit_log
from utils.logging import logging
from transformation.dm_file_preparation import *
import win32com.client as win32
import pythoncom
import pyautogui

def read_csv_skiprows(path):
  with open(path) as f:
    lines = f.readlines()
    num = [i for i, l in enumerate(lines) if l.startswith("No")]
    num = 0 if len(num) == 0 else num[0] - 1
    return num

def clean_file(onefile):
  df = pd.DataFrame(onefile.replace(np.nan, '', regex = True))
  df2 = pd.DataFrame(df.replace('nan', '', regex = False))
  if 'No.' in df2.columns:
    df3 = pd.DataFrame(df2.replace('NaT', '', regex = False))
    df3.dropna(subset=['No.'], inplace=True)
    df3['No.'] = df3['No.'].astype(int)
  elif 'No' in df2.columns:
    df3 = pd.DataFrame(df2.replace('NaT', '', regex = False))
    df3.dropna(subset=['No'], inplace=True)
    df3['No'] = df3['No'].astype(int)
  elif 'NO.' in df2.columns:
    df3 = pd.DataFrame(df2.replace('NaT', '', regex = False))
    df3.dropna(subset=['NO.'], inplace=True)
    df3['NO.'] = df3['NO.'].astype(int)
  elif 'NO' in df2.columns:
    df3 = pd.DataFrame(df2.replace('NaT', '', regex = False))
    df3.dropna(subset=['NO'], inplace=True)
    df3['NO'] = df3['NO'].astype(int) 
  else:
    df3 = pd.DataFrame(df2.replace('NaT', '', regex = False))
  return df3

def fixering(df, wb, sheets, fixer, sheet_num, keywerd):
  i = 0
  start_cell = search_excel(str(keywerd), sheet_num, wb)
  if start_cell != None:
    last_row = int(search_last_row(start_cell, fixer, sheet_num, wb, str(keywerd)) - 1)
  else:
    start_cell = search_excel(str(keywerd), sheet_num, wb)
    last_row = int(search_last_row(start_cell, fixer, sheet_num, wb, str(keywerd)) - 1)
  for i in range(len(df)):
    update_entry(wb.active, df.loc[df.index == i], last_row + i, i, start_column = 1)
  wb.save(str(fixer))
  wb.close()

def revising(fixer, sheetname, client):
  wb = load_workbook(fixer)
  wb.active = wb.worksheets.index(wb[sheetname])
  df = pd.read_excel(fixer, sheet_name = int(wb.worksheets.index(wb[sheetname])))
  dict, lst_str_cols2 = dict_str(df)
  searched_df = pd.read_excel(fixer, sheet_name = int(wb.worksheets.index(wb[sheetname])), dtype = dict)
  wb.close()
  return searched_df

def skiprows_input(source):
  wb = load_workbook(source)
  sheets = wb.sheetnames
  firstrow = ''
  kwrd = ''
  for v in range(len(sheets)):
    wb.active = wb.worksheets.index(wb[sheets[v]])
    sheet_num = int(wb.worksheets.index(wb[sheets[v]]))
    kwd = ['FirstColumn', 'No','NO','No.','NO.','POLICY APPROVAL DATE', 'Client Name', 'Member\'s Name', 'Policy Number']
    for i in range(len(kwd)):
      try:
        firstrow = int(search_excel(kwd[i], sheet_num, wb)[1]) - int(1)
        kwrd = kwd[i]
      except:
        wb.close()
        pass
      continue
  wb.close()
  return firstrow, kwrd

def skiprow_txt(source, seperator):
  skipr = [0, 1]
  for i in range(len(skipr)):
    try:
      onefile = pd.read_csv(source, sep = seperator, encoding = "ISO-8859-1", skiprows = skipr[i], header = None)
      skiprow = skipr[i]
      return skiprow
    except:
      pass
    continue

def skiprow_xls(df):
  firstrow = ''
  kwd = ['No','NO','No.','NO.', 'Client Name', 'Header', 'Company name', 'Import Type']
  for i in range(len(kwd)):
    try:
      firstrow = int(df[df[0]==kwd[i]].index.values)
      kwrd = kwd[i]
    except:
      pass
    continue
  return firstrow,kwrd


def find_sep(source):
  sep = ''
  # Open the file in read only mode
  with open(source, 'r', errors = 'ignore') as read_obj:
      # Read all lines in the file one by one
      for line in read_obj:
          # For each line, check if line contains the string
          try:
            if '|' in line:
              sep = '|'
            elif '^' in line:
              sep = '^'
            elif ';' in line:
              sep = ';'
            elif '>' in line:
              sep = '>'
            else:
              pass
          except:
            pass
  return sep

def dict_str(df):
  lst_str_cols = ['Address2', 'Address3', 'I.C.No', 'Hand Phone',
                 'House Phone', 'Work Phone', 'IC No (New)', 'NRIC ID',
                  'Postal Code', 'IC', 'ADDRESS03', 'ADDRESS02',
                  'Address 2', 'Address 3', 'NRIC', 'Principal NRIC', 'I.C. No',
                  'Phone', 'New IC No.', 'Agent Mobile', 'Postcode', 'OtherIC',
                  'External Ref Id (aka Client)', 'New IC no', 'Other Iden No',
                  'Other Iden Code', 'IC No(New)', 'New NRIC', '              X', 'I/C NUMBER ',
                  'Birth Cert/','IC No', 'New IC (new)', 'Unnamed: 10', 'I.C. No(New)',
                  'IC_NO', 'PRINCIPAL_IC', 'Principals Phone Number','Ic No(new)']
  lst_str_cols2 = list()
  for i in range(len(lst_str_cols)):
    try:
      if lst_str_cols[i] in df.columns:                                 
        lst_str_cols2.append(lst_str_cols[i])
      else:
        pass
    except:
      pass
    continue
  dict_dtypes = {x : 'str'  for x in lst_str_cols2}
  return dict_dtypes, lst_str_cols2

def changedtype(df4, lst_str_cols2):
  for w in range(len(lst_str_cols2)):
    try:
      if str('External Ref Id (aka Client)').lower() == str(lst_str_cols2[w].lower()):
        df4[lst_str_cols2[w]] = df4[lst_str_cols2[w]].str.replace('\.0', '', regex=True)
      elif str('OtherIC').lower() == str(lst_str_cols2[w].lower()):
        df4[lst_str_cols2[w]] = df4[lst_str_cols2[w]].str.replace('\.0', '', regex=True)
      else:
        df4[lst_str_cols2[w]] = df4[lst_str_cols2[w]].str.split(".")[0]
    except:
      pass
    continue
  return df4

def cleanup_df(df):
  df1 = pd.DataFrame(df.replace('-', '', regex = False))
  df1.dropna(axis = 0, how = 'all', thresh = None, subset = None, inplace = True)
  return df1

def date_validation(df):
  for i in df.columns:
    try:
      kwd = ['Date', 'DATE', 'DOB', 'date', 'Previous Policy End Date', 'Cancellation Date']
      for j in range(len(kwd)):
        try:
          if str(kwd[j]) in str(i):
            df[i] = df[i].dt.strftime('%d/%m/%Y')
        except:
          pass
        continue
    except:
      pass
    continue
  df1 = pd.DataFrame(df.replace(np.nan, '', regex = True))
  df2 = pd.DataFrame(df1.replace('nan', '', regex = False))
  df3 = pd.DataFrame(df2.replace('NaT', '', regex = False))
  return df3

def readxcelfile(file):
  try:
    if os.path.exists(file) == True:
      print("Source path exists")
      file2 = file
      pass
    else:
      print("Fixer path repath")
      file2 = str(file).replace(str('\\'), str('\\\\'))
    import win32com.client as win32, win32com
    import pythoncom
    import pyautogui
    pythoncom.CoInitialize()
    try:
      excel = win32.gencache.EnsureDispatch('Excel.Application')
    except:
      excel = win32com.client.DispatchEx("Excel.Application")
    wb = excel.Workbooks.Open(file2)
    excel.EnableEvents = False
    excel.DisplayAlerts = False
    excel.Visible = False
    wb.Save()
    wb.Close(True)
    excel.Quit()
    print("Successfully read workbook with Excel")
  except Exception as error:
    import win32com.client as win32, win32com
    import pythoncom
    import pyautogui
    pythoncom.CoInitialize()
    try:
      excel = win32.gencache.EnsureDispatch('Excel.Application')
    except:
      excel = win32com.client.DispatchEx("Excel.Application")
    wb = excel.Workbooks.Open(file2)
    excel.EnableEvents = False
    excel.DisplayAlerts = False
    excel.Visible = False
    wb.Save()
    wb.Close(True)
    excel.Quit()
    print("Failed to read workbook with Excel")
    pass

def read_file(source, filename, result, failed, client, fixer, prefix, stpdm_base):
  wb = load_workbook(fixer)
  sheets = wb.sheetnames
  firstrow = ''
  keywerd = ''
  firstrow_input = ''
  for v in range(len(sheets)):
    if 'raw' in str(sheets[v].lower()):
      wb.active = wb.worksheets.index(wb[sheets[v]])
      sheet_num = int(wb.worksheets.index(wb[sheets[v]]))
      kwd = ['No','NO','No.','NO.', 'Client Name', 'Header', 'Company name', 'Import Type', 'MSIG SEQUENCE', 'Member\'s Name', 'Policy Number', 'Running number']
      for i in range(len(kwd)):
        try:
          firstrow = int(search_excel(kwd[i], sheet_num, wb)[1]) - int(1)
          keywerd = kwd[i]
        except:
          pass
        continue
    try:
      if '.xlsx' in str(source.lower()):
        print("Load Fixer - .xlsx file")
        try:
          firstrow_input, kwrd = skiprows_input(source)
        except:
          print("Failed to determine the firstrow")
        df = pd.read_excel(source, skiprows = firstrow_input).astype(str)
        if df[kwrd][0] == '----':
          dict, lst_str_cols2 = dict_str(df)
          df1 = pd.read_excel(source, dtype = dict, skiprows = firstrow_input).astype(str)
          for i in df1.index:
            if df1[kwrd][i] == '----' or df1[kwrd][i] == 'nan':
              df1 = df1.drop(index = i)
          onefile = df1.reset_index(drop=True)
        else:
          dict, lst_str_cols2 = dict_str(df)
          onefile = pd.read_excel(source, dtype = dict, skiprows = firstrow_input).astype(str)
        col_tag = True
      elif '.xls' in str(source.lower()):
        print("Load Fixer - .xls file")
        df = pd.read_excel(source, header = None).astype(str)
        firstrow_xls, kwrd = skiprow_xls(df)
        df1 = pd.read_excel(source, skiprows = firstrow_xls).astype(str)
        if df1[kwrd][0] == 'nan'and df1[kwrd][1] == 'nan':
          df2 = pd.read_excel(source, skiprows = firstrow_xls+2).astype(str)
          dict, lst_str_cols2 = dict_str(df2)
          onefile = pd.read_excel(source, dtype = dict, skiprows = firstrow_xls+2).astype(str)
        else:
          dict, lst_str_cols2 = dict_str(df1)
          onefile = pd.read_excel(source, dtype = dict, skiprows = firstrow_xls).astype(str)
        col_tag = True
      elif '.csv' in str(source.lower()):
        print("Load Fixer - .csv file")
        try:
          df = pd.read_csv(source, encoding = "ISO-8859-1", skiprows = firstrow).astype(str)
          dict, lst_str_cols2 = dict_str(df)
          onefile = pd.read_csv(source, encoding = "ISO-8859-1", skiprows = firstrow, dtype = dict).astype(str)
          onefile = onefile.rename(columns = {onefile.columns[0]:keywerd})
          col_tag = True
        except:
          df = pd.read_csv(source, encoding = "ISO-8859-1", skiprows = firstrow).astype(str)
          dict, lst_str_cols2 = dict_str(df)
          onefile = pd.read_csv(source, encoding = "ISO-8859-1", skiprows = firstrow, dtype = dict).astype(str)
          onefile = onefile.rename(columns = {onefile.columns[0]:keywerd})
          col_tag = True
      elif '.txt' in str(source.lower()):
        print("Load Fixer - .txt file")
        try:
          seperator = find_sep(source)
          skiprowtxt = skiprow_txt(source, seperator)
          onefile = pd.read_csv(source, sep = seperator, skiprows = skiprowtxt, encoding = "ISO-8859-1", header = None, dtype = str, low_memory = False )
          onefile = onefile.rename(columns = {0:keywerd})
          col_tag = True
        except:
          seperator = find_sep(source)
          skiprowtxt = skiprow_txt(source, seperator)
          onefile = pd.read_csv(source, sep = seperator, skiprows = skiprowtxt, header = None, encoding = "ISO-8859-1", dtype = str, error_bad_lines = False, low_memory = False)
          #onefile = pd.read_csv(source, sep = seperator, skiprows = skiprowtxt, header = None, dtype = str)
          onefile = onefile.rename(columns = {0:keywerd})
          col_tag = True
        pass
    except Exception as error:
      print("Error in parsing columns")
      col_tag = False
      shutil.move(source, str(failed), stpdm_base)      
      logging("STP-DM - No columns to parse from file.", error, stpdm_base)
      pass
    continue

  if onefile.empty != True and col_tag == True:
    head, tail = get_file_name(source)
    audit_log("STP-DM - File [ %s ] is not empty and will be processed in Fixer." %tail, "Completed...", stpdm_base)
    try:
      df = clean_file(onefile)
      try:
        print('Start fixering')
        fixering(df, wb, sheets, fixer, sheet_num, keywerd)
        print('Done fixering')
      except Exception as error:
        print("Error in FIXERing")
        logging("STP-DM - Error in FIXERing.", error, stpdm_base)
        pass

      #readxcelfile(fixer)

      try:
        if os.path.exists(fixer) == True:
          print("Fixer path exists")
          fixer2 = fixer
          pass
        else:
          print("Fixer path repath")
          fixer2 = str(fixer).replace(str('\\'), str('\\\\'))
        import win32com.client as win32, win32com
        import pythoncom
        import pyautogui
        pythoncom.CoInitialize()
        try:
          excel = win32.gencache.EnsureDispatch('Excel.Application')
        except:
          excel = win32com.client.DispatchEx("Excel.Application")
        wb = excel.Workbooks.Open(fixer2)
        excel.EnableEvents = False
        excel.DisplayAlerts = False
        excel.Visible = False
        wb.Save()
        wb.Close(True)
        excel.Quit()
        print("Successfully read workbook with Excel")
      except Exception as error:
        import win32com.client as win32, win32com
        import pythoncom
        import pyautogui
        pythoncom.CoInitialize()
        try:
          excel = win32.gencache.EnsureDispatch('Excel.Application')
        except:
          excel = win32com.client.DispatchEx("Excel.Application")
        wb = excel.Workbooks.Open(fixer2)
        excel.EnableEvents = False
        excel.DisplayAlerts = False
        excel.Visible = False
        wb.Save()
        wb.Close(True)
        excel.Quit()
        print("Failed to read workbook with Excel")
        print("Error: %s" %error)
        logging("STP-DM - Error in opening in-memory FIXER file.", error, stpdm_base)
        pass

      try:
        if str("_IO_") in prefix:
          prefix_ip = prefix.replace("_IO_", "_IP_")
          prefix_op = prefix.replace("_IO_", "_OP_")
          for m in range(len(sheets)):
            if 'MARC' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC', client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              dict, lst_str_cols2 = dict_str(df2)
              df2 = changedtype(df2, lst_str_cols2)
              df2.to_excel(result + "\\" + prefix + filename + ".xlsx", sheet_name = 'MARC', index = False)
              with pd.ExcelWriter(result + "\\" + prefix + filename + ".xlsx", engine = 'openpyxl', mode = 'a') as writer:
                for m in range(len(sheets)):
                  if 'CARD' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m]), client)
                    df.to_excel(writer, sheet_name = str(sheets[m]), index = False)
                  elif 'RAW' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m].upper()), client)
                    df.to_excel(writer, sheet_name = str(sheets[m].upper()), index = False)
                writer.save()
            elif 'MARC (OP)' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC (OP)',client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              dict, lst_str_cols2 = dict_str(df2)
              df2 = changedtype(df2, lst_str_cols2)
              df2.to_excel(result + "\\" + prefix_op + filename + ".xlsx", sheet_name = 'MARC (OP)', index = False)
              with pd.ExcelWriter(result + "\\" + prefix_op + filename + ".xlsx", engine = 'openpyxl', mode = 'a') as writer:
                for m in range(len(sheets)):
                  if 'CARD' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m]), client)
                    df.to_excel(writer, sheet_name = str(sheets[m]), index = False)
                  elif 'RAW' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m].upper()), client)
                    df.to_excel(writer, sheet_name = str(sheets[m].upper()), index = False)
                writer.save()
            elif 'MARC (IP)' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC (IP)', client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              dict, lst_str_cols2 = dict_str(df2)
              df2 = changedtype(df2, lst_str_cols2)
              df2.to_excel(result + "\\" + prefix_ip + filename + ".xlsx", sheet_name = 'MARC (IP)', index = False)
              with pd.ExcelWriter(result + "\\" + prefix_ip + filename + ".xlsx", engine = 'openpyxl', mode = 'a') as writer:
                for m in range(len(sheets)):
                  if 'CARD' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m]), client)
                    df.to_excel(writer, sheet_name = str(sheets[m]), index = False)
                  elif 'RAW' in str(sheets[m].upper()):
                    df = revising(fixer, str(sheets[m].upper()), client)
                    df.to_excel(writer, sheet_name = str(sheets[m].upper()), index = False)
                writer.save()
        else:
          for m in range(len(sheets)):
            if 'MARC' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC',client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              dict, lst_str_cols2 = dict_str(df2)
              df2 = changedtype(df2, lst_str_cols2)
              dict, lst_str_cols2 = dict_str(df2)
              df2 = changedtype(df2, lst_str_cols2)
              df2.to_excel(result + "\\" + prefix + filename + ".xlsx", sheet_name = 'MARC', index = False)
            elif 'MARC (OP)' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC (OP)',client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              dict, lst_str_cols2 = dict_str(df2)
              df2 = changedtype(df2, lst_str_cols2)
              df2.to_excel(result + "\\" + prefix + filename + ".xlsx", sheet_name = 'MARC (OP)', index = False)
            elif 'MARC (IP)' == str(sheets[m].upper()):
              df = revising(fixer, 'MARC (IP)', client)
              df1 = cleanup_df(df)
              df2 = date_validation(df1)
              dict, lst_str_cols2 = dict_str(df2)
              df2 = changedtype(df2, lst_str_cols2)
              df2.to_excel(result + "\\" + prefix + filename + ".xlsx", sheet_name = 'MARC (IP)', index = False)
          with pd.ExcelWriter(result + "\\" + prefix + filename + ".xlsx", engine = 'openpyxl', mode = 'a') as writer:
            for m in range(len(sheets)):
              if 'CARD' in str(sheets[m].upper()):
                df = revising(fixer, str(sheets[m]),client)
                df.to_excel(writer, sheet_name = str(sheets[m]), index = False)
              elif 'RAW' in str(sheets[m].upper()):
                df = revising(fixer, str(sheets[m].upper()),client)
                df.to_excel(writer, sheet_name = str(sheets[m].upper()), index = False)
            writer.save()

      except Exception as error:
        print("Error in Revising dataframe from FIXER")
        logging("STP-DM - Error in Revising dataframe from FIXER.", error, stpdm_base)
        pass

    except Exception as error:
      shutil.move(source, failed, stpdm_base)
      print("STP-DM - Error in file cleaning.")
      logging("STP-DM - Error in file cleaning.", error, stpdm_base)
      pass

  elif onefile.empty == True:
    shutil.move(source, failed, stpdm_base)

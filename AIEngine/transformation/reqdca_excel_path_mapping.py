#!/usr/bin/python
# FINAL SCRIPT updated as of 8th July 2020
# Workflow - REQ/DCA

# Declare Python libraries needed for this script
import glob
import openpyxl
import pandas as pd
import json
import time
from datetime import datetime as dt
from openpyxl import load_workbook
from utils.audit_trail import audit_log
from connector.connector import MySqlConnector
from transformation.reqdca_excel_path_mapping import *
from transformation.reqdca_excel_manipulation import *

# Find if there's any new file located in master file folders
def find_new_file_location(filelocation):

  # Load json file containing all path
  try:
    #with open(r'C:\Users\adm.alfred.simbun\OneDrive - AXA\ALFRED - Development\TESTING\REQDCA_Paths\Testing_Available_File.json') as f:
    with open(r'\\Dtisvr2\cba_uat\Result\REQDCA_Paths\Testing_Available_File.json') as f:
      data = json.load(f)

    # find the unavailable file
    filelocation = list(set(filelocation) - set(data))
    data.extend(filelocation)
    #with open(r'C:\Users\adm.alfred.simbun\OneDrive - AXA\ALFRED - Development\TESTING\REQDCA_Paths\Testing_Available_File.json', 'w') as f:
    with open(r'\\Dtisvr2\cba_uat\Result\REQDCA_Paths\Testing_Available_File.json', 'w') as f:
      json.dump(data, f)
    return filelocation
  except Exception as error:
    print("Path Library is up-to-date")
    print(error)

# Retrieve all Adm No available in each Excel File
def get_case_ID(filelocation):

  keyword = 'Adm. No'
  df1 = pd.DataFrame()
  #unprocess_df = pd.DataFrame()

  for i in range(len(filelocation)):
    try:
      print(filelocation[i])
      wb = load_workbook(filelocation[i])
      sheetname = wb.sheetnames
      for j in range(len(sheetname)):
        cell_no = search_excel(keyword,j, wb)
        df = pd.read_excel(filelocation[i], header = cell_no[1]-1, sheet_name = j)
        main_df = df.loc[:,keyword]
        main_df = main_df.dropna().reset_index().drop(columns = "index")
        main_df = main_df.astype(str)

        if main_df.empty == False:
          try:
            new = main_df[keyword].str.split("-", n = 1, expand = True)
            main_df['Case ID 2'] = new[0]
            main_df['Subcase ID'] = new[1]
          except Exception as error:
            main_df['Case ID 2'] = main_df[keyword]
            main_df['Subcase ID'] = None
          Filename = [filelocation[i]]*(main_df.shape[0])
          main_df['Filename'] = Filename
          df1 = df1.append(main_df, ignore_index = True)
        else:
          continue
    except Exception as error:
      print(error)
      print("%s CANNOT be processed. Please check." % filelocation[i])
      continue
  return df1

# Perform query on Marc Database to find the necessary information based on Adm No
def query_SQL(df):
  try:
    conn = MySqlConnector()
    cur = conn.cursor()

    for i in range(df.shape[0]):
      print(i)
      try:
        if(df.iloc[i]['Subcase ID'] == None):
          cur.execute("""SELECT
          cse.`inptCaseId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmClientName`, inpt.`inptCaseMmPlanName`,
          inpt.`inptCaseMmInsurerName`, inpt.`inptCaseMMPrgName`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmExtRefId`
          FROM inpt_case cse
          INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
          INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
          WHERE inptCaseId = '%s' LIMIT 1""" % df.iloc[i]['Case ID 2'])
          fetched_result = cur.fetchall()
        else:
          cur.execute("""SELECT
          cse.`inptCaseId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmClientName`, inpt.`inptCaseMmPlanName`,
          inpt.`inptCaseMmInsurerName`, inpt.`inptCaseMMPrgName`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmExtRefId`
          FROM inpt_case cse
          INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
          INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
          INNER JOIN `inpt_case_csu` subcase ON subcase.`inptCasePpInptCaseId` = cse.`inptCaseId`
          WHERE inptCaseId = '%s' AND subcase.`inptCasePpId` = '%s'""" % (df.iloc[i]['Case ID 2'], df.iloc[i]['Subcase ID']))
          fetched_result = cur.fetchall()

        df.loc[i,'Pol Num'] = fetched_result[0][1]
        df.loc[i,'Client'] = fetched_result[0][2]
        df.loc[i,'Bill To'] = fetched_result[0][4]
        df.loc[i,'Ext Ref'] = fetched_result[0][7]
      except Exception as error:
        continue
    cur.close()
    conn.close()
    return df
  except Exception as error:
    print(error)

# Create a mapping file in jason based on Ext Ref Num 
def create_map(df):
  try:
    #path_df = r'C:\Users\adm.alfred.simbun\OneDrive - AXA\ALFRED - Development\TESTING\REQDCA_Paths\Testing_Path_DataFrame.json'
    path_df = r'\\Dtisvr2\cba_uat\Result\REQDCA_Paths\Testing_Path_DataFrame.json'
    with open(path_df, 'r') as f:
      path_df = json.load(f)
      path_df = pd.DataFrame.from_dict(path_df, orient = 'index')

    for i in range(df.shape[0]):
      print(i)
      path_df = path_df.append({'Year':2019,'Pol Num': df.iloc[i]['Pol Num'],'Ext Ref': df.iloc[i]['Ext Ref'],'Client': df.iloc[i]['Client'],'Bill To': df.iloc[i]['Bill To'],'Path': df.iloc[i]['Filename']}, ignore_index = True)

    path_df = path_df.sort_values(by = ['Path']).drop_duplicates(subset = ['Year','Ext Ref'], keep = 'first').reset_index().drop(columns = 'index')
    path_df.to_json(r'\\Dtisvr2\cba_uat\Result\REQDCA_Paths\Testing_Path_DataFrame.json', orient = 'index')

  except Exception as error:
    print(error)

# Perform Excel Mapping process
def map_excel_file():
  try:
    now = time.strftime("%d-%m-%Y")
    year = dt.now().year

    # Find all current year Excel file in Master File
    print("- Set the file path")
    #path = r"C:\Users\adm.alfred.simbun\OneDrive - AXA\ALFRED - Development\TESTING\MASTER\\"
    #path = r"\\fs1\finance\1-CREDIT CONTROL\2-MASTER & OPS INVOICES\MASTER\\"
    path = r"\\Dtisvr2\cba_uat\Result\REQDCA_Paths\MASTER\\"
    filelocation = [f for f in glob.glob(path + "**/*" + str(year) + ".xlsx", recursive = True)]

    print("- Register the paths")
    filelocation_unique = find_new_file_location(filelocation)

    print("- Get the path based on Case ID")
    if filelocation_unique == []:
      df = []
      print("- Filepath has the latest updated info")
    else:
      df = pd.DataFrame(get_case_ID(filelocation_unique))
      print("- Peform SQL query")
      df = query_SQL(df)
      print("- Create file paths mapping")
      create_map(df)

    return df

  except Exception as error:
    print(error)

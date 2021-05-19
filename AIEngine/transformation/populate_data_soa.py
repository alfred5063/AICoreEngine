#!/usr/bin/python
# FINAL SCRIPT updated as of 1st April 2021
# Workflow - Finance SOA
# Version 1

# Declare Python libraries needed for this script
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import xlrd
from datetime import datetime
from utils.Session import session
from connector.dbconfig import read_db_config
from utils.audit_trail import audit_log
from utils.logging import logging
from connector.connector import MySqlConnector
from transformation.reqdca_excel_manipulation import *
from automagic.automagica_finance_soa import *
from transformation.soa_mapping import *
from transformation.soa_client_mapping import *
from transformation.soa_check_due_date import *



def vendor_transaction(vt_excel, base):
  print('Begin Appending and Populating New Column in Vendor Transaction')
  try:
    vt_df = pd.DataFrame(pd.read_excel(vt_excel, header = None))
    #copy column d + e to the last column
    vt_df[11] = vt_df[3]
    vt_df[12] = vt_df[4]
    df = pd.DataFrame(vt_df.values, columns = ['Invoice No', 'Doc Types', 'Client', 'GL No', 'Register Date', 'Due Date', 'Batch-Entry', 'Paid Invoice No/Payment Ref',
                                               'Invoice Amount', 'Paid Amount', 'Balance', 'Payment Batch No', 'Payment Date'])
    df.to_excel(vt_excel, index = False)
    print('Completed... New Column Created')
  except Exception as error:
    logging('Vendor Transaction Excel Append Process', error, base)
    print(error)

def populate_data_SOA(soa_excel, vt_excel, clear_date_excel, base):
  audit_log('Hospital SOA File Mapping', 'Begin Mapping SOA File', base)
  print('Begin Mapping SOA File')
  key = 'Invoice ID'

  try:
    # Load SOA Workbook as Dataframe
    soa_workbook = load_workbook(soa_excel)
    soa_sheets = soa_workbook.worksheets
    soa_df = pd.read_excel(soa_excel, sheet_name = soa_sheets[0].title, skiprows = 4, usecols = list(range(soa_sheets[0].max_column+1)))
    
    # Mapping SOA columns
    soa_ori_df = pd.read_excel(soa_excel, sheet_name = soa_sheets[0].title, skiprows = 4, usecols = list(range(soa_sheets[0].max_column+1)))
    soa_df.columns = soa_mapping(soa_df, base)
    soa_df = soa_df.rename(columns = {"Invoice No": "Invoice ID"})

    audit_log('Finance SOA - Hospital SOA File Mapping', 'Completed...', base)
    soa_df = soa_df.dropna(axis = 0, how = 'any', subset = [key])

    # Store existing header in SOA df as list
    dropped_header = soa_df.columns.tolist()
    invoice_list = ['Invoice No', 'Invoice ID', 'Bill No', 'Document No', 'Inv', 'Ref No', 'Bill Reference', 'A/C No.', 'Description', 'Doc. No', 'Visit No']
    key = list(set(dropped_header) & set(invoice_list))[0]
    
    # Create additional column to be appended into SOA Excel
    additional_column = ['Overdue Day', 'GL Ref', 'Client', 'Invoice Amount', 'Register Date', 'Payment Date', 'Payment Batch No', 'Paid Invoice No/Payment Ref', 'Paid Amount', 'Bank Clear Date', 'Insurance', 'Submission Date', 'Bank-in Date', 'Amount (RM)', 'Remarks']
    new_column = pd.DataFrame(additional_column)
    new_column = pd.DataFrame.transpose(new_column)

    # Load VT Workbook as Dataframe
    vt_workbook = load_workbook(vt_excel)
    vt_sheets = vt_workbook.worksheets
    vt_df = pd.read_excel(vt_excel, sheet_name = vt_sheets[0].title, skiprows = 0, usecols = list(range(vt_sheets[0].max_column+1)))

    # Filter VT Dataframe according to IN Doc Types
    temp_vt_df = vt_df[vt_df['Doc Types'] == 'IN']
    temp_vt_df_PY = vt_df[vt_df['Batch-Entry'] == 'PY']
    temp_vt_df_PY = temp_vt_df_PY[temp_vt_df_PY['Paid Invoice No/Payment Ref'].str.contains('PY', case = False)]
    temp_vt_df_PY['Payment Batch No'] = temp_vt_df_PY['Payment Batch No'].astype(int)
    temp_vt_df_PY = temp_vt_df_PY[['Payment Batch No', 'Paid Invoice No/Payment Ref']]

    # Load Bank Clear Date Workbook
    clear_date_workbook = load_workbook(clear_date_excel)
    clear_date_sheets = clear_date_workbook.worksheets
    clear_date_df = pd.read_excel(clear_date_excel, sheet_name = clear_date_sheets[0].title, skiprows = 0, usecols = list(range(clear_date_sheets[0].max_column+1)))
    # Perform merging between SOA and VT Dataframe
    soa_df = pd.merge(soa_df, temp_vt_df[['Invoice No', 'GL No', 'Client', 'Invoice Amount', 'Register Date']], how = 'left', left_on = [key], right_on = ['Invoice No']).drop(['Invoice No'], axis = 1)
    # Add new column and add value 'Mediclinic Case' if GL No is null and GL Ref No contains LG and SP, 'International' if GL No is null and GL Ref No contains 'KUL'
    soa_df['Remarks'] = np.nan
    x = 0
    for x in range(len(soa_df['GL No_x'])):
      if 'SP' in str(soa_df['GL No_x'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to Mediclinic-Please refer to Adilah'
      elif 'SP' in str(soa_df['GL No_y'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to Mediclinic-Please refer to Adilah'
      elif 'LG' in str(soa_df['GL No_x'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to Mediclinic-Please refer to Adilah'
      elif 'LG' in str(soa_df['GL No_y'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to Mediclinic-Please refer to Adilah'
      elif 'OP' in str(soa_df['GL No_x'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to Mediclinic-Please refer to Adilah'
      elif 'OP' in str(soa_df['GL No_y'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to Mediclinic-Please refer to Adilah'
      elif 'KUL' in str(soa_df['GL No_x'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to International-Please refer to Syaiful/Farrah'
      elif 'KUL' in str(soa_df['GL No_y'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to International-Please refer to Syaiful/Farrah'
      elif 'kul' in str(soa_df['GL No_x'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to International-Please refer to Syaiful/Farrah'
      elif 'kul' in str(soa_df['GL No_y'][x]):
        soa_df.loc[x, 'Remarks'] = 'Belong to International-Please refer to Syaiful/Farrah'
      else:
        pass
    # Rename GL No_x to GL No so MARC can process
    soa_df = soa_df.rename(columns = {'GL No_x' : 'GL No'})
    soa_df = soa_df.rename(columns = {'GL No_y' : 'GL Ref'})

    # Accessing MARC
    soa_df = query_Marc(soa_df, base)
    soa_df['Remarks'] = np.where(pd.notnull(soa_df['GL Ref']) & soa_df['Client'].str.lower().str.contains("ge life"), 'GE Life Case - Please refer to GE Life for payment', soa_df['Remarks'])
    soa_df = pd.merge(soa_df, vt_df[['Paid Invoice No/Payment Ref', 'Payment Date', 'Payment Batch No']], how = 'left', left_on = [key], right_on = ['Paid Invoice No/Payment Ref'])
    soa_df = pd.merge(soa_df, vt_df[['GL No', 'Paid Invoice No/Payment Ref']], how = 'left', left_on = [key], right_on = ['GL No'], suffixes = ('', '_y')).drop(['GL No_y'], axis = 1)
    soa_df = pd.merge(soa_df, vt_df[['Paid Invoice No/Payment Ref','Paid Amount']], how = 'left', left_on = [key], right_on = ['Paid Invoice No/Payment Ref'], suffixes = ('', '_y')).drop(['Paid Invoice No/Payment Ref_y'], axis=1)

    # Change type of a column into int as there is issue with mapping process
    clear_date_df['Payment Batch No'] = clear_date_df['Payment Batch No'].astype(int)
    soa_df['Payment Batch No'] = soa_df['Payment Batch No'].fillna(0).astype(int)
    test_df = soa_df.merge(clear_date_df, left_on = 'Payment Batch No', right_on = 'Payment Batch No').merge(temp_vt_df_PY, left_on = 'Payment Batch No', right_on = 'Payment Batch No')
    test_df['Payment Batch No'] = np.where(test_df['Payment Batch No'] == 0, '', test_df['Payment Batch No'].astype(str).str.zfill(12))
    test_df['Paid Invoice No/Payment Ref_x'] = test_df['Paid Invoice No/Payment Ref_y']
    test_df = test_df.rename(columns = {'Paid Invoice No/Payment Ref_x' : 'Paid Invoice No/Payment Ref', 'Client_x' : 'Client'})

    # Drop duplicate value in SOA Dataframe
    key = 'Invoice ID'
    test_df = test_df.drop_duplicates(subset = [key, 'Payment Date', 'Paid Amount'], keep = 'first', inplace = False).reset_index(drop = True)
    soa_df = soa_df.drop_duplicates(subset = [key, 'GL Ref'], keep = 'first', inplace = False).reset_index(drop = True)
    pd.DataFrame.transpose(test_df)

    # Replace the correct data
    a = 0
    for a in range(soa_df.shape[0]):
      b = 0
      for b in range(test_df.shape[0]):
        if soa_df.iloc[a]['Invoice ID'] == test_df.iloc[b]['Invoice ID']:
          soa_df.loc[a, 'Paid Invoice No/Payment Ref'] = test_df.iloc[b]['Paid Invoice No/Payment Ref']
          soa_df.loc[a, 'Payment Date'] = test_df.iloc[b]['Payment Date']
          soa_df.loc[a, 'Payment Batch No'] = test_df.iloc[b]['Payment Batch No']
          soa_df.loc[a, 'Paid Amount'] = test_df.iloc[b]['Paid Amount']
          soa_df.loc[a, 'Bank Clear Date'] = test_df.iloc[b]['Bank Clear Date']
        else:
          pass

      if soa_df.iloc[a]['Payment Batch No'] == 0:
        soa_df.loc[a, 'Payment Batch No'] = np.nan
        soa_df.loc[a, 'Paid Invoice No/Payment Ref'] = np.nan
        soa_df.loc[a, 'Payment Date'] = np.nan
        soa_df.loc[a, 'Paid Amount'] = np.nan
        soa_df.loc[a, 'Bank Clear Date'] = np.nan
      elif 'PY' in soa_df.iloc[a]['Paid Invoice No/Payment Ref']:
        pass
      else:
        soa_df.loc[a, 'Payment Batch No'] = np.nan
        soa_df.loc[a, 'Paid Invoice No/Payment Ref'] = np.nan
        soa_df.loc[a, 'Payment Date'] = np.nan
        soa_df.loc[a, 'Paid Amount'] = np.nan
        soa_df.loc[a, 'Bank Clear Date'] = np.nan

    first_column_name = soa_ori_df.columns[0]
    last_length = len(soa_ori_df.columns) - 1
    last_column_name = soa_ori_df.columns[last_length]
    x = 0
    y = len(soa_ori_df.columns)
    for x in range(len(soa_ori_df.columns)):
      if x <= last_length:
        if soa_df.columns[x] != soa_ori_df.columns[x]:
          test = soa_ori_df.columns[x]
          soa_df.columns = [test.replace(soa_df.columns[x], soa_ori_df.columns[x]) for test in soa_df]
        else:
          pass
      else:
        pass

    # Rename the dropped_header columns name same with soa_df
    first_column_name = soa_ori_df.columns[0]
    last_length = len(soa_ori_df.columns) - 1
    last_column_name = soa_ori_df.columns[last_length]
    x = 0
    for x in range(len(soa_ori_df.columns)):
      if x <= last_length:
        if dropped_header[x] != soa_ori_df.columns[x]:
          test = soa_ori_df.columns[x]
          dropped_header = [test.replace(dropped_header[x], soa_ori_df.columns[x]) for test in dropped_header]
        else:
          pass
      else:
        pass

    gl_no_backup = soa_df['GL Ref']
    invoice_amount_back_up = soa_df['Invoice Amount_y']

    column_format = dropped_header + additional_column
    soa_df = soa_df.reindex(column_format, axis = 1)


    # Repopulate data into the right column
    soa_df['GL Ref'] = gl_no_backup
    soa_df['Invoice Amount'] = invoice_amount_back_up

    # Rename the soa_df columns name back to ori columns name
    check_list = list(soa_df.columns)
    if 'Outstanding' in check_list:
      soa_df['Outstanding'] = soa_df['Outstanding'].round(2)
    else:
      pass

    audit_log('Hospital SOA File Mapping', 'Completed...', base)
    print('Completed... SOA File Maped')

  except Exception as error:
    logging('Hospital SOA File Mapping', error, base)
    print(error)

  audit_log('Hospital SOA File Update', 'Begin Updating the Hospital SOA File', base)
  print('Hospital SOA File Update')
  # Client Master File Mapping
  soa_df = edited_cmf(soa_df, base)
  # Check due date
  soa_df = check_due_date(soa_df, vt_df)
  # New remarks requirement to be add
  i = 0
  for i in range(soa_df.shape[0]):
    if pd.notnull(soa_df.iloc[i]['GL Ref']) and pd.notnull(soa_df.iloc[i]['Client']) and pd.notnull(soa_df.iloc[i]['Invoice Amount']) and pd.notnull(soa_df.iloc[i]['Register Date']) and pd.notnull(soa_df.iloc[i]['Payment Date']) and pd.notnull(soa_df.iloc[i]['Payment Batch No']) and pd.notnull(soa_df.iloc[i]['Paid Invoice No/Payment Ref']) and pd.notnull(soa_df.iloc[i]['Paid Amount']) and pd.notnull(soa_df.iloc[i]['Bank Clear Date']) and pd.isnull(soa_df.iloc[i]['Remarks']):
      soa_df.loc[i, 'Remarks'] = 'PAID. Please offset'
    # If GL Ref - Paid Amount not empty buy Bank clear date is empty - Payment issued. Pending approval
    elif pd.notnull(soa_df.iloc[i]['GL Ref']) and pd.notnull(soa_df.iloc[i]['Client']) and pd.notnull(soa_df.iloc[i]['Invoice Amount']) and pd.notnull(soa_df.iloc[i]['Register Date']) and pd.notnull(soa_df.iloc[i]['Payment Date']) and pd.notnull(soa_df.iloc[i]['Payment Batch No']) and pd.notnull(soa_df.iloc[i]['Paid Invoice No/Payment Ref']) and pd.notnull(soa_df.iloc[i]['Paid Amount']) and pd.isnull(soa_df.iloc[i]['Bank Clear Date']) and pd.isnull(soa_df.iloc[i]['Remarks']):
      soa_df.loc[i, 'Remarks'] = 'Payment issued. Pending approval'
    elif pd.isnull(soa_df.iloc[i]['GL Ref']) and pd.isnull(soa_df.iloc[i]['Client']) and pd.isnull(soa_df.iloc[i]['Invoice Amount']) and pd.isnull(soa_df.iloc[i]['Register Date']) and pd.isnull(soa_df.iloc[i]['Remarks']):
      soa_df.loc[i, 'Remarks'] = 'Bill not received. Please provide CTC'
    elif pd.notnull(soa_df.iloc[i]['GL Ref']) and pd.notnull(soa_df.iloc[i]['Client']) and pd.notnull(soa_df.iloc[i]['Invoice Amount']) and pd.notnull(soa_df.iloc[i]['Register Date']) and pd.isnull(soa_df.iloc[i]['Payment Date']) and pd.isnull(soa_df.iloc[i]['Payment Batch No']) and pd.isnull(soa_df.iloc[i]['Paid Invoice No/Payment Ref']) and pd.isnull(soa_df.iloc[i]['Paid Amount']) and pd.isnull(soa_df.iloc[i]['Bank Clear Date']) and pd.isnull(soa_df.iloc[i]['Remarks']):
      soa_df.loc[i, 'Remarks'] = 'Not due yet'
    else:
      pass
  try:
    print('#Append dataframe to SOA Excel')
    start_cell = 6
    for i in range(soa_df.shape[0]):
      update_entry(soa_sheets[0], soa_df, start_cell, i, 1)
      start_cell = start_cell + 1

    print('#Append new header to SOA Excel')
    start_append = len(dropped_header) + 1
    update_entry(soa_sheets[0], new_column, 5, 0, start_append)

    save_and_close(soa_workbook, soa_excel)
    audit_log('Hospital SOA File Update', 'Completed...', base)
    
    print('Completed... the Hospital SOA File is Updated')

  except Exception as error:
    logging('Hospital SOA File Update', error, base)
    print(error)


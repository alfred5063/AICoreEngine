#!/usr/bin/python
# FINAL SCRIPT updated as of 20th July 2020
# Workflow - REQ/DCA

# Declare Python libraries needed for this script
import openpyxl
import time
from datetime import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from transformation.reqdca_excel_manipulation import *
from connector.connector import MySqlConnector
from utils.logging import logging
from utils.audit_trail import audit_log
from directory.files_listing_operation import *
from directory.get_filename_from_path import *
from directory.createfolder import *

# Determine the current year
current_year = dt.now().year
current_date = time.strftime("%Y-%m-%d")

def update_dcm(newmain_df, document, doc_template, email, DCMfile, reqdca_base):

  try:

    # Query MARC database to get user's abbreviated name
    conn = MySqlConnector()
    cur = conn.cursor()

    # Query in MARC database
    # Take note that querying development databse might not give the correct result.
    paramater = [str(email),]
    stored_proc = cur.callproc('dca_query_marc_for_user_email', paramater)
    for i in cur.stored_results():
      results = i.fetchall()
    if results != []:
      shortname = results[0][0]
    else:
      shortname = ''

    # Close the current connection
    cur.close()
    conn.close()

    # Keyword to search the cells in Excel to be used as header in the dataframe
    if document == 'DCA':
      keyword1 = 'Date'
      keyword2 = 'Reason for adjustment'
    else:
      keyword1 = 'Date'
      keyword2 = 'Initial'

    # Perform update data into new row in Disbursement Claim Master file
    i = 0
    for i in range(len(newmain_df)):

      a = 0
      # Load Disbursement Claim Master File
      wb = load_workbook(DCMfile)
      sheetname = wb.sheetnames

      # Locate the index of the active sheet
      sheet = wb.index(wb.active)
      activesheet = wb.active

      filename = get_file_name(DCMfile)
      if(os.path.isdir(str(filename[0]) + "\\TEMPORARY\\") == False):
        createfolder(str(filename[0]) + "\\TEMPORARY\\", reqdca_base, parents = True, exist_ok = True)
      if(os.path.isdir(str(filename[0]) + "\\BACKUP\\") == False):
        createfolder(str(filename[0]) + "\\BACKUP\\", reqdca_base, parents = True, exist_ok = True)
      myfile = list(getListOfFiles(str(filename[0]) + "\\TEMPORARY\\"))
      mylist = pd.DataFrame(list(filter(lambda x: str(filename[1]) in x, myfile)), columns = ['Filenames'])
      if mylist.empty == True:
        shutil.copy(DCMfile, str(filename[0]) + "\\BACKUP\\")
        shutil.move(DCMfile, str(filename[0]) + "\\TEMPORARY\\")
        myfile = list(getListOfFiles(str(filename[0]) + "\\TEMPORARY\\"))
        mylist = pd.DataFrame(list(filter(lambda x: str(filename[1]) in x, myfile)), columns = ['Filenames'])
        filelocation = mylist['Filenames'][0]
      else:
        filelocation = mylist['Filenames'][0]

      # Search for cell number of the keyword in Excel sheet
      start_cell = search_excel(keyword1, sheet, wb)
      end_cell = search_excel(keyword2, sheet, wb)

      try:
        # Create empty dataframe with header
        dcm_df = pd.DataFrame(pd.read_excel(filelocation, sheet_name = sheet, skiprows = start_cell[1]-1, nrows = 0))
        if document == 'DCA':
          dcm_df = dcm_df.rename(columns = {"Types of invoice ": "Types of invoice"})
          newmain_df2 = newmain_df[newmain_df['Adm. No'].astype(str).str.contains("%s" % newmain_df['Adm. No'][i])]
          dcm_df2 = update_dataframe_dca(dcm_df, newmain_df2, a, shortname, doc_template, reqdca_base)
        else:
          newmain_df2 = newmain_df[newmain_df['Case ID'].astype(str).str.contains("%s" % newmain_df['Case ID'][i])]
          dcm_df2 = update_dataframe_req(dcm_df, newmain_df2, a, shortname, doc_template, reqdca_base)

        last_row = search_last_row(start_cell, filelocation, sheet, wb, keyword1)
        start_row = last_row + 1
        update_entry(activesheet, dcm_df2, start_row, a, start_column = 1)
        wb.save(str(DCMfile))
        os.remove(filelocation)
      except Exception as error:
        audit_log("REQ/DCA - Failed to create and update DCM file.", "Completed...", reqdca_base)
        logging("REQ/DCA - Failed to create and update DCM file.", error, reqdca_base)
        pass

    wb.close()

  except Exception as error:
    logging("REQDCA - reqdca_update_dcm", error, reqdca_base)
    pass

# DCA - Update the data in dataframe from various sources
def update_dataframe_dca(dcm_df, newmain_df2, a, shortname, doc_template, reqdca_base):

  try:
    dcm_df.loc[a,'Date'] = time.strftime("%d/%m/%Y")
    dcm_df.loc[a,'Types of invoice'] = doc_template #document
    dcm_df.loc[a,'Dis. claims no'] = newmain_df2.iloc[a]['Disbursement Claims'] #H&S latest running number
    dcm_df.loc[a,'No. of cases'] = '1' #fixed

    if newmain_df2.iloc[a]['Type_x'] == 'Admission':
      dcm_df.loc[a,'File no'] = newmain_df2.iloc[a]['Adm. No']
      dcm_df.loc[a,'Claims listing no'] = str(newmain_df2.iloc[a]['Disbursement Listing'])
    else:
      dcm_df.loc[a,'File no'] = str(newmain_df2.iloc[a]['Adm. No']) + "-" + str(newmain_df2.iloc[a]['Sub Case ID_x'])
      dcm_df.loc[a,'Claims listing no'] = str(newmain_df2.iloc[a]['Disbursement Listing'])

    dcm_df.loc[a,'Customer Name_Master'] = newmain_df2.iloc[a]['Client Name']
    dcm_df.loc[a,'Hospital'] = str(newmain_df2.iloc[a]['Hospital Name'])
    dcm_df.loc[a,'Bill no.'] = newmain_df2.iloc[a]['Bill. Num.']
    dcm_df.loc[a,'Patient'] = newmain_df2.iloc[a]['''Member's Name''']
    dcm_df.loc[a,'Amounts'] = str(newmain_df2.iloc[a]['Total'])
    dcm_df.loc[a,'Action'] = ""
    dcm_df.loc[a,'Reasons'] = newmain_df2.iloc[a]['Type_x'] # Pick-list from the user
    dcm_df.loc[a,'Initial'] = shortname
    dcm_df.loc[a,'Bank in date'] = ""
    dcm_df.loc[a,'Cheque no / TT'] = ""
    dcm_df.loc[a,'Amount Paid'] = ""
    dcm_df.loc[a,'Remarks'] = newmain_df2.iloc[a]['Reason']
    dcm_df.loc[a,'Reason for adjustment'] = newmain_df2.iloc[a]['Remarks_x'] # Pick-list from the user

    return dcm_df
  except Exception as error:
    logging("REQDCA - reqdca_update_dcm", error, reqdca_base)

# REQ - Update the data in dataframe from various sources
def update_dataframe_req(dcm_df, newmain_df2, a, shortname, doc_template, reqdca_base):

  try:
    dcm_df.loc[a,'Date'] = time.strftime("%d/%m/%Y")
    dcm_df.loc[a,'Disbursement claims no'] = newmain_df2.iloc[a]['HNS Number'] #H&S latest running number
    dcm_df.loc[a,'SAGE ID'] = str(newmain_df2.iloc[a]['Case ID']) + "-REQ"

    if newmain_df2.iloc[a]['Type'] == 'Admission':
      dcm_df.loc[a,'File'] = newmain_df2.iloc[a]['Case ID']
      mycln = newmain_df2.iloc[a]['Client Listing Number'].split(" ")
      if str(mycln[0]) != 'N/A':
        dcm_df.loc[a,'Claims listing no'] = str(mycln[0]) + " (C)"
      else:
        dcm_df.loc[a,'Claims listing no'] = str(mycln[0])
    else:
      dcm_df.loc[a,'File'] = str(newmain_df2.iloc[a]['Case ID']) + "-" + str(newmain_df2.iloc[a]['Sub Case ID'])
      mycln = newmain_df2.iloc[a]['Client Listing Number'].split(" ")
      if str(mycln[0]) != 'N/A':
        dcm_df.loc[a,'Claims listing no'] = str(mycln[0]) + " (C)"
      else:
        dcm_df.loc[a,'Claims listing no'] = str(mycln[0])

    dcm_df.loc[a,'Bill to'] = newmain_df2.iloc[a]['Insurer Name']
    dcm_df.loc[a,'Client'] = newmain_df2.iloc[a]['Client Name']
    dcm_df.loc[a,'Patient Name'] = newmain_df2.iloc[a]['Patient Name']
    dcm_df.loc[a,'Status (VIP / Non-Vip Tan chong claims) '] = ""
    dcm_df.loc[a,'Policy no '] = newmain_df2.iloc[a]['Insurence Policy Number']
    dcm_df.loc[a,'Admission date'] = newmain_df2.iloc[a]['Admission Date']
    dcm_df.loc[a,'Discharge date'] = newmain_df2.iloc[a]['Discharged Date']
    dcm_df.loc[a,'Hospital'] = str(newmain_df2.iloc[a]['Hospital Name'])
    dcm_df.loc[a,'Bill no'] = str(newmain_df2.iloc[a]['Bill. Num.'])
    dcm_df.loc[a,'Amounts'] = str(newmain_df2.iloc[a]['Amount'])
    dcm_df.loc[a,'OB Received date'] = newmain_df2.iloc[a]['OB Received Date']
    dcm_df.loc[a,'OB Registered date'] = newmain_df2.iloc[a]['OB Registered Date']
    dcm_df.loc[a,'Reasons'] = newmain_df2.iloc[a]['Reason'] #Free text from the user
    dcm_df.loc[a,'Cashless / Post / Fruit Basket / Reimbursement '] = newmain_df2.iloc[a]['Remarks'] #pick list from the user (Cashless / Post / Fruit Basket / Reimbursement)
    dcm_df.loc[a,'Initial'] = shortname

    return dcm_df
  except Exception as error:
    logging("REQDCA - reqdca_update_dcm", error, reqdca_base)

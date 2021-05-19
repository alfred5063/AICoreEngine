#!/usr/bin/python
# FINAL SCRIPT updated as of 8th July 2020
# Workflow - REQ/DCA Update CML File

# Declare Python libraries needed for this script
import os
import json
import time
import glob
from datetime import datetime as dt
import pandas as pd
from directory.files_listing_operation import getListOfFiles
from loading.excel.checkExcelLoading import *
from transformation.reqdca_excel_manipulation import *
import openpyxl
from zipfile import BadZipfile

# Get filename based on client name
def search_cmlfile(dcamain_df, cmlfiles, current_year):

  myfile = list(getListOfFiles(cmlfiles))

  i = 0
  temp_df2 = pd.DataFrame()
  finalpath_df = pd.DataFrame()
  finalpath_df2 = pd.DataFrame()
  clientlist = pd.DataFrame()
  for i in range(len(dcamain_df['Client Name'])):
    client = dcamain_df['Client Name'][i].lower().split(" ")
    filelist = list(filter(lambda x: '%s' %str(current_year) in x, myfile))
    filelist = pd.DataFrame(list(filter(lambda x: '.xlsx' in x, filelist)), columns = ['Filelist'])

    a = 0
    for a in range(len(client)):
      temp_df = pd.DataFrame(columns = ['Filelist'])
      searched = filelist[filelist['Filelist'].str.lower().astype(str).str.contains(client[a])]
      if searched.empty != True:
        temp_df = temp_df.append(searched)
        temp_df = temp_df.drop_duplicates().reset_index().drop(columns = ['index'])
        for b in range(len(temp_df['Filelist'])):
          temp_df.loc[b, 'Case ID'] = str(dcamain_df['Case ID'][i])
          temp_df.loc[b, 'Sub Case ID'] = str(dcamain_df['Sub Case ID'][i])
          temp_df.loc[b, 'Type'] = str(dcamain_df['Type'][i])
        filelist = searched
        temp_df2 = temp_df2.append(temp_df)
        finalpath_df = temp_df2.drop_duplicates()
      else:
        pass
      continue

    finalpath_df2 = finalpath_df2.append(finalpath_df)
  finalpath_df2 = finalpath_df2.drop_duplicates().reset_index().drop(columns = ['index'])
  return finalpath_df2


# Search for Adm.No in specific file and update it
def update_cmf(dcamain_df, current_year, cmlfiles):

  # Get the client path list
  myclient = search_cmlfile(dcamain_df, cmlfiles, current_year)
  dcamain_df['Path'] = ''

  class GetOutOfLoop(Exception):
    pass

  f = 0
  for f in range(len(dcamain_df['Case ID'])):
    try:
      if dcamain_df['Type'][f] == 'POST':
        mycaseid = dcamain_df['Sub Case ID'][f]
        queried = myclient[myclient['Sub Case ID'].astype(str).str.contains("%s" % int(mycaseid))]
      else:
        mycaseid = dcamain_df['Case ID'][f]
        queried = myclient[myclient['Case ID'].astype(str).str.contains("%s" % int(mycaseid))]

      queried = queried.drop_duplicates().reset_index().drop(columns = ['index'])

      i = 0
      for i in range(len(queried)):
        filelocation = queried['Filelist'][i]
        filestatus = is_locked(filelocation)

        if str(filestatus) == 'False':
          first_header = 'Adm. No'

          try:
            #wb = load_workbook(filelocation, read_only = True, keep_vba = False, data_only = True, guess_types = False, keep_links = False)
            wb = load_workbook(filelocation, read_only = True, keep_vba = False, data_only = True, keep_links = False)
          except BadZipfile:
            print("BAD FILE")
            break;

          sheets = wb.sheetnames
          a = 0
          for a in range(len(sheets)):
            start_cell = search_excel(first_header, a, wb)
            if start_cell != None:
              searched_df = pd.read_excel(filelocation, sheet_name = a, skiprows = start_cell[1]-1)
              searched_df = searched_df[pd.notnull(searched_df['Adm. No'])]
              searched_df['Adm. No'] = searched_df['Adm. No'].astype(str)
              searched_df = searched_df[searched_df['Adm. No'].astype(str).str.contains("%s" % str(mycaseid))]
              if searched_df.empty != True:
                dcamain_df.loc[f, 'Path'] = filelocation
                raise GetOutOfLoop
              else:
                dcamain_df.loc[f, 'Path'] = 'Path not found'
                pass
            else:
              pass
        else:
          raise GetOutOfLoop
    except GetOutOfLoop:
      pass
    continue
  return dcamain_df




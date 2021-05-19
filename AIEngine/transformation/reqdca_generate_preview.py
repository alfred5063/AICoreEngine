#!/usr/bin/python
# FINAL SCRIPT updated as of 20th July 2020
# Workflow - REQ/DCA

# Declare Python libraries needed for this script
import json
import csv
import pandas as pd
import re
import numpy as np
import ast
import os, time
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from transformation.reqdca_excel_manipulation import *
from datetime import datetime as dt
from connector.connector import MySqlConnector, MSSqlConnector
from utils.audit_trail import audit_log
from utils.logging import logging
import pyodbc as db
from loading.query.query_as_df import *

# Function to prepare
def extract_info(jsondf_val, reqdca_base, document, DCMfile):

  try:

    # Current year
    curyear = time.strftime("%y", time.localtime())
    current_year = dt.now().year
    #for testing
    #curyear = '20'
    #current_year = '2020'

    # Create temporary empty dataframe
    temp_df = pd.DataFrame()
    insurer_results = pd.DataFrame()
    invalid_df = pd.DataFrame(columns = ['Case ID', 'Sub Case ID', 'Member Ref. ID', 'Client Listing Number', 'Patient Name', 'Amount', 'Reason', 'Bill. Num.',
                                         'Hospital Name', 'Insurence Plan Name', 'Insurence Policy Number', 'Admission Date', 'Client Name', 'Insurer Name',
                                         'Attention To', 'Address', 'HNS Number', 'Type', 'Client ID', 'Remarks', 'Validity'])

    # create a cursor for MARC database connection
    conn_mysql = MySqlConnector()
    cur_mysql = conn_mysql.cursor()
    connector = MSSqlConnector

    # Query MSSQL for Insurer's Address
    query = '''SELECT atten_1, comp_name, comp_address, identifier, id FROM cba.insurer_details'''
    fetched_clientdetail = mssql_get_df_by_query_without_param(query, reqdca_base, connector)
    i = 0
    for i in range(len(fetched_clientdetail['identifier'])):
      fetched_clientdetail.loc[i, 'identifier'] = str(fetched_clientdetail.loc[i, 'identifier']).rstrip('\r\n')
      #print(fetched_clientdetail.loc[i, 'identifier'])
    # Query Insurer's Name
    c = 0
    for c in range(len(jsondf_val)):
      caseid = int(jsondf_val.iloc[c]['Case ID'])
      type = str(jsondf_val.iloc[c]['Type'])
      parameters = [str(caseid),]

      if type == 'POST':

        # Calling Stored Procedure
        #stored_proc = cur_mysql.callproc('query_marc_for_outpatient_validity', parameters)
        #for z in cur_mysql.stored_results():
        #  results = z.fetchall()

        #for testing
        z = 0
        query_list = list()
        for z in range(len(parameters)):
          query = '''SELECT DISTINCT
                    cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
                    inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`,
                    inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`, subcase.`inptCasePpId`
                    FROM inpt_case cse
                    INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
                    INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
                    INNER JOIN `inpt_case_csu` subcase ON subcase.`inptCasePpInptCaseId` = cse.`inptCaseId`
                    INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
                    INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
                    WHERE subcase.`inptCasePpId` = {0}'''.format(parameters[z])
          query_list.append(query)

        i = 0
        results = list()
        fetched_result = pd.DataFrame()
        for i in range(len(query_list)):
          cur_mysql.execute(query_list[i])
          result = cur_mysql.fetchall()
          results.append(result)
          fetched_result = fetched_result.append(results[i])

        #End of testing code

        #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
        #for j in cur_mysql.stored_results():
        #  results_cln = j.fetchall()

        #for testing 2
        j = 0
        query_list2 = list()
        for j in range(len(parameters)):
          query = '''SELECT DISTINCT acc.`inptCaseAccInptCaseId`, acc.`inptCaseAccChqNum`
                    FROM inpt_case_accounting acc
                    WHERE acc.inptCaseAccType = "BD" AND acc.`inptCaseAccInptCaseId` = {0}'''.format(parameters[j])
          query_list2.append(query)

        i = 0
        results_cln = pd.DataFrame()
        for i in range(len(query_list2)):
          cur_mysql.execute(query_list2[i])
          result = cur_mysql.fetchall()
          results_cln = results_cln.append(result)

        #End of testing code

        #fetched_result = pd.DataFrame(results)
        #results_cln = pd.DataFrame(results_cln)
        fetched_result = fetched_result.drop_duplicates()
        results_cln = results_cln.drop_duplicates()
        fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg','CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                   'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName',  'CaseMmClientName',
                                   'CaseMmInsurerName', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate', 'CasePpId']
        results_cln.columns = ['Sub Case ID', 'Client Listing Number']
        if results_cln.empty != True:
          fetched_result['CaseAccChqNum'] = results_cln['Client Listing Number']
        else:
          fetched_result['CaseAccChqNum'] = "N/A"
      else:

        #stored_proc = cur_mysql.callproc('query_marc_for_inpatient_validity', parameters)
        #for z in cur_mysql.stored_results():
        #  results = z.fetchall()

        #for testing
        z = 0
        query_list = list()
        for z in range(len(parameters)):
          query = '''SELECT DISTINCT cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
          inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, acc.`inptCaseAccChqNum`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`, inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`
          FROM inpt_case cse
          INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
          INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
          INNER JOIN `inpt_case_accounting` acc ON acc.`inptCaseAccInptCaseId` = cse.`inptCaseId`
          INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
          INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
          WHERE acc.inptCaseAccType = "BD" AND cse.`inptCaseId` = {0}'''.format(parameters[z])
          query_list.append(query)

        i = 0
        results = list()
        fetched_result = pd.DataFrame()
        for i in range(len(query_list)):
          cur_mysql.execute(query_list[i])
          result = cur_mysql.fetchall()
          results.append(result)
          fetched_result = fetched_result.append(results[i])


        #End of testing code

        #fetched_result = pd.DataFrame(results)
        pd.set_option('max_columns', None)
        fetched_result = fetched_result.drop_duplicates()
        fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                   'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName', 'CaseMmClientName',
                                   'CaseMmInsurerName', 'CaseAccChqNum', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate']
        fetched_result['CasePpId'] = "0"

      param = [str(fetched_result['CasemmPolicyNum'][0]), str(fetched_result['CaseMmPlanExtCode'][0]), str(fetched_result['CaseMmPlanName'][0])]

      #stored_proc = cur_mysql.callproc('dca_query_marc_for_insurer_details', param)
      #for z in cur_mysql.stored_results():
      #  results = z.fetchall()

      #for testing 3
      z = 0
      query_list3 = list()
      #for j in range(len(param)):
      query = '''SELECT DISTINCT inptCaseMmPlanName, inptCaseMmPolicyNum, inptCaseMmPlanExtCode, inptCaseMmPolType, inptCaseMmPrgName, inptCaseMmPrgSvcId, inptCaseMmInsurerName
                FROM inpt_case_member
                WHERE inptCaseMmPolicyNum IN ('{0}') AND inptCaseMmPlanExtCode IN ('{1}')'''.format(param[0], param[1])
      query_list3.append(query)

      i = 0
      myquery = pd.DataFrame()
      for i in range(len(query_list3)):
        cur_mysql.execute(query_list3[i])
        result = cur_mysql.fetchall()
        insurer_results = myquery.append(result)

      #End of testing code




      #myquery = pd.DataFrame(results)
      #insurer_results = insurer_results.append(myquery)

    insurer_results = insurer_results.drop_duplicates()
    insurer_results.columns = ['CaseMmPlanName', 'CaseMmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPolType', 'CaseMmPrgName', 'CaseMmPrgSvcId', 'CaseMmInsurerName']
    
    # Merge both data
    insurer_results['CaseMmInsurerName'] = insurer_results['CaseMmInsurerName'].str.lower()
    fetched_clientdetail['comp_name'] = fetched_clientdetail['comp_name'].str.lower()
    merged_insurer_details = pd.merge(left = insurer_results, right = fetched_clientdetail, left_on = 'CaseMmInsurerName', right_on = 'comp_name')
    merged_insurer_details = merged_insurer_details.drop_duplicates()

    if document == 'DCA':
      # DCA - Query the latest HSN number from MSSQL database
      query = '''SELECT content FROM cba.disbursement_master WHERE year = %s'''
      params = (current_year)
      fetched_hns = mssql_get_df_by_query(query, params, reqdca_base, connector)
      fetched_hns = fetched_hns.to_json()
      fetched_hns = ast.literal_eval(fetched_hns)
      fetched_hns = fetched_hns['content']['0']
      fetched_hns = pd.DataFrame(json.loads(fetched_hns))
      fetched_hns_curr = fetched_hns.values[-1].tolist()
      fetched_hns_curr = fetched_hns_curr[0]['Disbursement Claims No'].split("-")
      hsn_curyear = fetched_hns_curr[0].split("H&S")
      if hsn_curyear[1] != curyear:
        fetched_hns_curr[0] = "H&S" + curyear
        fetched_hns_curr[1] = "0".zfill(5)
      else:
        fetched_hns_curr[0] = fetched_hns_curr[0]
        fetched_hns_curr[1] = fetched_hns_curr[1]
      n = 1
      while (fetched_hns_curr[0] == ''):
        fetched_hns_curr = fetched_hns.values[-n].tolist()
        fetched_hns_curr = fetched_hns_curr[0]['Disbursement Claims No'].split("-")
        if fetched_hns_curr[0] == '':
          n = n + 1
        else:
          a = 0
          b = 1
          hsn_curyear = fetched_hns_curr[a].split("H&S")

          HNS = fetched_hns_curr[a] + "-" + str(int(fetched_hns_curr[b]) + 1).zfill(5)
      else:
        a = 0
        b = 1
        HNS = fetched_hns_curr[a] + "-" + str(int(fetched_hns_curr[b]) + 1).zfill(5)

      # Get the latest REQ number from REQ file
      wb = load_workbook(DCMfile)
      sheetname = wb.sheetnames

      # Locate the index of the active sheet
      sheet = wb.index(wb.active)
      activesheet = wb.active
      keyword1 = 'Date'
      start_cell = search_excel(keyword1, sheet, wb)
      dcm_df = pd.DataFrame(pd.read_excel(DCMfile, sheet_name = sheet, skiprows = start_cell[1]-1))
      searched_df = dcm_df[pd.notnull(dcm_df['Date'])]
      latest_hnsfromxls = str(max(searched_df['Dis. claims no'])).split("-")
      hsnxls_curyear = latest_hnsfromxls[0].split("H&S")
      if hsnxls_curyear[1] != curyear:
        latest_hnsfromxls[0] = "H&S" + curyear
        latest_hnsfromxls[1] = "0".zfill(5)
      else:
        latest_hnsfromxls[0] = latest_hnsfromxls[0]
        latest_hnsfromxls[1] = latest_hnsfromxls[1]
      HNSxls = latest_hnsfromxls[0] + "-" + str(int(latest_hnsfromxls[1]) + 1).zfill(5)

      if HNSxls != HNS:
        HNS = HNSxls
      else:
        HNS = HNSxls

      wb.close()

    else:

      # REQ - Query the latest HSN number from MSSQL database
      query = '''SELECT content FROM cba.ops_disbursement_req_master WHERE year = %s'''
      params = (current_year)
      fetched_hns = mssql_get_df_by_query(query, params, reqdca_base, connector)
      fetched_hns = fetched_hns.to_json()
      fetched_hns = ast.literal_eval(fetched_hns)
      fetched_hns = fetched_hns['content']['0']
      fetched_hns = pd.DataFrame(json.loads(fetched_hns))
      fetched_hns_curr = fetched_hns.values[-1].tolist()
      fetched_hns_curr = fetched_hns_curr[0]['Disbursement claims no'].split("-")
      hsn_curyear = fetched_hns_curr[0].split("REQ")
      if hsn_curyear[1] != curyear:
        fetched_hns_curr[0] = "REQ" + curyear
        fetched_hns_curr[1] = "0".zfill(4)
      else:
        fetched_hns_curr[0] = fetched_hns_curr[0]
        fetched_hns_curr[1] = fetched_hns_curr[1]
      n = 1
      while (fetched_hns_curr[0] == ''):
        fetched_hns_curr = fetched_hns.values[-n].tolist()
        fetched_hns_curr = fetched_hns_curr[0]['Disbursement claims no'].split("-")
        if fetched_hns_curr[0] == '':
          n = n + 1
        else:
          a = 0
          b = 1
          HNS = fetched_hns_curr[a] + "-" + str(int(fetched_hns_curr[b]) + 1).zfill(3)
      else:
        a = 0
        b = 1
        HNS = fetched_hns_curr[a] + "-" + str(int(fetched_hns_curr[b]) + 1).zfill(3)
        print(HNS)
      # Get the latest REQ number from REQ file
      wb = load_workbook(DCMfile)
      sheetname = wb.sheetnames
      print(sheetname)
      # Locate the index of the active sheet
      sheet = wb.index(wb.active)
      activesheet = wb.active
      keyword1 = 'Date'
      start_cell = search_excel(keyword1, sheet, wb)
      dcm_df = pd.DataFrame(pd.read_excel(DCMfile, sheet_name = sheet, skiprows = start_cell[1]-1))
      searched_df = dcm_df[pd.notnull(dcm_df['Date'])]
      latest_hnsfromxls = str(max(searched_df['Disbursement claims no'])).split("-")
      hsnxls_curyear = latest_hnsfromxls[0].split("REQ")
      if hsnxls_curyear[1] != curyear:
        latest_hnsfromxls[0] = "REQ" + curyear
        latest_hnsfromxls[1] = "0".zfill(4)
      else:
        latest_hnsfromxls[0] = latest_hnsfromxls[0]
        latest_hnsfromxls[1] = latest_hnsfromxls[1]
      HNSxls = latest_hnsfromxls[0] + "-" + str(int(latest_hnsfromxls[1]) + 1).zfill(3)

      if HNSxls != HNS:
        HNS = HNSxls
      else:
        HNS = HNSxls

      wb.close()

    # Extract out valid Case ID
    jsondf_valid = jsondf_val.loc[jsondf_val['Validity'] == "Valid Case ID"].set_index('Case ID').reset_index()
    if jsondf_valid.empty == True:
      jsondf_invalid = jsondf_val.loc[jsondf_val['Validity'] != "Valid Case ID"].set_index('Case ID').reset_index()
      v = 0
      invalidid = jsondf_invalid['Case ID']
      for i in range(len(invalidid)):
        invalid_df.loc[i, 'Case ID'] = "Case ID [ %s ] is invalid." % invalidid[i]
        invalid_df.loc[i, 'Amount'] = int(jsondf_invalid['Amount'][i])
        invalid_df.loc[i, 'Remarks'] = jsondf_invalid['Remarks'][i]
        invalid_df.loc[i, 'Type'] = jsondf_invalid['Type'][i]
        invalid_df.loc[i, 'Reason'] = jsondf_invalid['Reason'][i]
        invalid_df.loc[i, 'Validity'] = jsondf_invalid['Validity'][i]

      # Consolidating temp_df with invalid_df to get a finalized dataframe
      temp_df = temp_df.append(invalid_df, ignore_index = True, sort = False)
      temp_df = temp_df.drop_duplicates()
      temp_df = pd.DataFrame(temp_df)

    else:
      # Perform some SQL quesries and table joining
      i = 0
      mycaseid = jsondf_valid['Case ID']
      for i in range(len(mycaseid)):
        caseid = int(jsondf_valid.iloc[i]['Case ID'])
        type = str(jsondf_valid.iloc[i]['Type'])
        client = str(jsondf_valid.iloc[i]['Client'])
        arrangement = str(jsondf_valid.iloc[i]['Arrangement'])

        HNS = HNS
        parameters = [str(caseid),]

        if document == 'DCA':
          if type == 'POST':

            # Calling Stored Procedure
            #stored_proc = cur_mysql.callproc('query_marc_for_outpatient_validity', parameters)
            #for z in cur_mysql.stored_results():
            #  results = z.fetchall()

            #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
            #for j in cur_mysql.stored_results():
            #  results_cln = j.fetchall()

            #fetched_result = pd.DataFrame(results)
            #results_cln = pd.DataFrame(results_cln)

            #for testing
            z = 0
            query_list = list()
            for z in range(len(parameters)):
              query = '''SELECT DISTINCT
                        cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
                        inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`,
                        inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`, subcase.`inptCasePpId`
                        FROM inpt_case cse
                        INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
                        INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
                        INNER JOIN `inpt_case_csu` subcase ON subcase.`inptCasePpInptCaseId` = cse.`inptCaseId`
                        INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
                        INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
                        WHERE subcase.`inptCasePpId` = {0}'''.format(parameters[z])
              query_list.append(query)

            i = 0
            results = list()
            fetched_result = pd.DataFrame()
            for i in range(len(query_list)):
              cur_mysql.execute(query_list[i])
              result = cur_mysql.fetchall()
              results.append(result)
              fetched_result = fetched_result.append(results[i])


            #End of testing code

            #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
            #for j in cur_mysql.stored_results():
            #  results_cln = j.fetchall()

            #for testing 2
            j = 0
            query_list2 = list()
            for j in range(len(parameters)):
              query = '''SELECT DISTINCT acc.`inptCaseAccInptCaseId`, acc.`inptCaseAccChqNum`
                        FROM inpt_case_accounting acc
                        WHERE acc.inptCaseAccType = "BD" AND acc.`inptCaseAccInptCaseId` = {0}'''.format(parameters[j])
              query_list2.append(query)

            i = 0
            results_cln = pd.DataFrame()
            for i in range(len(query_list2)):
              cur_mysql.execute(query_list2[i])
              result = cur_mysql.fetchall()
              results_cln = results_cln.append(result)

            #End of testing code

            fetched_result = fetched_result.drop_duplicates()
            results_cln = results_cln.drop_duplicates()
            fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg','CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                       'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName',  'CaseMmClientName',
                                       'CaseMmInsurerName', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate', 'CasePpId']
            results_cln.columns = ['Sub Case ID', 'Client Listing Number']
            if results_cln.empty != True:
              fetched_result['CaseAccChqNum'] = results_cln['Client Listing Number']
            else:
              fetched_result['CaseAccChqNum'] = "N/A"

          else:

            #stored_proc = cur_mysql.callproc('query_marc_for_inpatient_validity', parameters)
            #for z in cur_mysql.stored_results():
            #  results = z.fetchall()
            #for testing
            z = 0
            query_list = list()
            for z in range(len(parameters)):
              query = '''SELECT DISTINCT cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
              inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, acc.`inptCaseAccChqNum`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`, inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`
              FROM inpt_case cse
              INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
              INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
              INNER JOIN `inpt_case_accounting` acc ON acc.`inptCaseAccInptCaseId` = cse.`inptCaseId`
              INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
              INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
              WHERE acc.inptCaseAccType = "BD" AND cse.`inptCaseId` = {0}'''.format(parameters[z])
              query_list.append(query)

            i = 0
            results = list()
            fetched_result = pd.DataFrame()
            for i in range(len(query_list)):
              cur_mysql.execute(query_list[i])
              result = cur_mysql.fetchall()
              results.append(result)
              fetched_result = fetched_result.append(results[i])


            #End of testing code
            #fetched_result = pd.DataFrame(results)
            fetched_result = fetched_result.drop_duplicates()
            fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                       'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName', 'CaseMmClientName',
                                       'CaseMmInsurerName', 'CaseAccChqNum', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate']
            fetched_result['CasePpId'] = "0"

        else:
          if client == ('Tan Chong Motor Holdings Berhad' or 'AXA Affin General Insurance Berhad') and type == 'POST':

            # Calling Stored Procedure
            #stored_proc = cur_mysql.callproc('query_marc_for_outpatient_validity', parameters)
            #for z in cur_mysql.stored_results():
            #  results = z.fetchall()

            #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
            #for j in cur_mysql.stored_results():
            #  results_cln = j.fetchall()

            #fetched_result = pd.DataFrame(results)
            #results_cln = pd.DataFrame(results_cln)

            #for testing
            z = 0
            query_list = list()
            for z in range(len(parameters)):
              query = '''SELECT DISTINCT
                        cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
                        inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`,
                        inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`, subcase.`inptCasePpId`
                        FROM inpt_case cse
                        INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
                        INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
                        INNER JOIN `inpt_case_csu` subcase ON subcase.`inptCasePpInptCaseId` = cse.`inptCaseId`
                        INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
                        INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
                        WHERE subcase.`inptCasePpId` = {0}'''.format(parameters[z])
              query_list.append(query)

            i = 0
            results = list()
            fetched_result = pd.DataFrame()
            for i in range(len(query_list)):
              cur_mysql.execute(query_list[i])
              result = cur_mysql.fetchall()
              results.append(result)
              fetched_result = fetched_result.append(results[i])

            #End of testing code

            #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
            #for j in cur_mysql.stored_results():
            #  results_cln = j.fetchall()

            #for testing 2
            j = 0
            query_list2 = list()
            for j in range(len(parameters)):
              query = '''SELECT DISTINCT acc.`inptCaseAccInptCaseId`, acc.`inptCaseAccChqNum`
                        FROM inpt_case_accounting acc
                        WHERE acc.inptCaseAccType = "BD" AND acc.`inptCaseAccInptCaseId` = {0}'''.format(parameters[j])
              query_list2.append(query)
      
            i = 0
            results_cln = pd.DataFrame()
            for i in range(len(query_list2)):
              cur_mysql.execute(query_list2[i])
              result = cur_mysql.fetchall()
              results_cln = results_cln.append(result)

            #End of testing code

            fetched_result = fetched_result.drop_duplicates()
            results_cln = results_cln.drop_duplicates()
            fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg','CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                       'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName',  'CaseMmClientName',
                                       'CaseMmInsurerName', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate', 'CasePpId']
            results_cln.columns = ['Sub Case ID', 'Client Listing Number']
            if results_cln.empty != True:
              fetched_result['CaseAccChqNum'] = results_cln['Client Listing Number']
            else:
              fetched_result['CaseAccChqNum'] = "N/A"

          elif client == ('Tan Chong Motor Holdings Berhad' or 'AXA Affin General Insurance Berhad') and type == 'Admission':

            #stored_proc = cur_mysql.callproc('query_marc_for_inpatient_validity', parameters)
            #for z in cur_mysql.stored_results():
            #  results = z.fetchall()

            #for testing
            z = 0
            query_list = list()
            for z in range(len(parameters)):
              query = '''SELECT DISTINCT cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
              inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, acc.`inptCaseAccChqNum`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`, inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`
              FROM inpt_case cse
              INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
              INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
              INNER JOIN `inpt_case_accounting` acc ON acc.`inptCaseAccInptCaseId` = cse.`inptCaseId`
              INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
              INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
              WHERE acc.inptCaseAccType = "BD" AND cse.`inptCaseId` = {0}'''.format(parameters[z])
              query_list.append(query)

            i = 0
            results = list()
            fetched_result = pd.DataFrame()
            for i in range(len(query_list)):
              cur_mysql.execute(query_list[i])
              result = cur_mysql.fetchall()
              results.append(result)
              fetched_result = fetched_result.append(results[i])

            #End of testing code
            #fetched_result = pd.DataFrame(results)
            fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                       'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName', 'CaseMmClientName',
                                       'CaseMmInsurerName', 'CaseAccChqNum', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate']
            fetched_result['CasePpId'] = "0"

          elif client == ('Allianz General Insurance Company (M) Bhd' or 'lonpac' in client.lower() or 'ytl' in client.lower()) and arrangement == 'Yes' and type == 'POST':

            #stored_proc = cur_mysql.callproc('query_marc_for_outpatient_validity', parameters)
            #for z in cur_mysql.stored_results():
            #  results = z.fetchall()

            #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
            #for j in cur_mysql.stored_results():
            #  results_cln = j.fetchall()

            #fetched_result = pd.DataFrame(results)
            #results_cln = pd.DataFrame(results_cln)

            #for testing
            z = 0
            query_list = list()
            for z in range(len(parameters)):
              query = '''SELECT DISTINCT
                        cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
                        inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`,
                        inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`, subcase.`inptCasePpId`
                        FROM inpt_case cse
                        INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
                        INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
                        INNER JOIN `inpt_case_csu` subcase ON subcase.`inptCasePpInptCaseId` = cse.`inptCaseId`
                        INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
                        INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
                        WHERE subcase.`inptCasePpId` = {0}'''.format(parameters[z])
              query_list.append(query)

            i = 0
            results = list()
            fetched_result = pd.DataFrame()
            for i in range(len(query_list)):
              cur_mysql.execute(query_list[i])
              result = cur_mysql.fetchall()
              results.append(result)
              fetched_result = fetched_result.append(results[i])

            #End of testing code

            #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
            #for j in cur_mysql.stored_results():
            #  results_cln = j.fetchall()

            #for testing 2
            j = 0
            query_list2 = list()
            for j in range(len(parameters)):
              query = '''SELECT DISTINCT acc.`inptCaseAccInptCaseId`, acc.`inptCaseAccChqNum`
                        FROM inpt_case_accounting acc
                        WHERE acc.inptCaseAccType = "BD" AND acc.`inptCaseAccInptCaseId` = {0}'''.format(parameters[j])
              query_list2.append(query)
      
            i = 0
            results_cln = pd.DataFrame()
            for i in range(len(query_list2)):
              cur_mysql.execute(query_list2[i])
              result = cur_mysql.fetchall()
              results_cln = results_cln.append(result)

            #End of testing code
            fetched_result = fetched_result.drop_duplicates()
            results_cln = results_cln.drop_duplicates()
            fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                       'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName',  'CaseMmClientName',
                                       'CaseMmInsurerName', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate', 'CasePpId']
            results_cln.columns = ['Sub Case ID', 'Client Listing Number']
            if results_cln.empty != True:
              fetched_result['CaseAccChqNum'] = results_cln['Client Listing Number']
            else:
              fetched_result['CaseAccChqNum'] = "N/A"

          elif client == ('Allianz General Insurance Company (M) Bhd' or 'lonpac' in client.lower() or 'ytl' in client.lower()) and arrangement == 'Yes' and type == 'Admission':

            #stored_proc = cur_mysql.callproc('query_marc_for_inpatient_validity', parameters)
            #for z in cur_mysql.stored_results():
            #  results = z.fetchall()

            #for testing
            z = 0
            query_list = list()
            for z in range(len(parameters)):
              query = '''SELECT DISTINCT cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
              inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, acc.`inptCaseAccChqNum`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`, inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`
              FROM inpt_case cse
              INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
              INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
              INNER JOIN `inpt_case_accounting` acc ON acc.`inptCaseAccInptCaseId` = cse.`inptCaseId`
              INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
              INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
              WHERE acc.inptCaseAccType = "BD" AND cse.`inptCaseId` = {0}'''.format(parameters[z])
              query_list.append(query)

            i = 0
            results = list()
            fetched_result = pd.DataFrame()
            for i in range(len(query_list)):
              cur_mysql.execute(query_list[i])
              result = cur_mysql.fetchall()
              results.append(result)
              fetched_result = fetched_result.append(results[i])

            #End of testing code

            #fetched_result = pd.DataFrame(results)
            fetched_result = fetched_result.drop_duplicates()
            fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                       'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName', 'CaseMmClientName',
                                       'CaseMmInsurerName', 'CaseAccChqNum', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate']
            fetched_result['CasePpId'] = "0"
          else:
            if type == 'POST':

              #stored_proc = cur_mysql.callproc('query_marc_for_outpatient_validity', parameters)
              #for z in cur_mysql.stored_results():
              #  results = z.fetchall()

              #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
              #for j in cur_mysql.stored_results():
              #  results_cln = j.fetchall()

              #fetched_result = pd.DataFrame(results)
              #results_cln = pd.DataFrame(results_cln)

              #for testing
              z = 0
              query_list = list()
              for z in range(len(parameters)):
                query = '''SELECT DISTINCT
                          cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
                          inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`,
                          inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`, subcase.`inptCasePpId`
                          FROM inpt_case cse
                          INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
                          INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
                          INNER JOIN `inpt_case_csu` subcase ON subcase.`inptCasePpInptCaseId` = cse.`inptCaseId`
                          INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
                          INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
                          WHERE subcase.`inptCasePpId` = {0}'''.format(parameters[z])
                query_list.append(query)

              i = 0
              results = list()
              fetched_result = pd.DataFrame()
              for i in range(len(query_list)):
                cur_mysql.execute(query_list[i])
                result = cur_mysql.fetchall()
                results.append(result)
                fetched_result = fetched_result.append(results[i])

              #End of testing code

              #stored_proc = cur_mysql.callproc('query_marc_for_cln', parameters)
              #for j in cur_mysql.stored_results():
              #  results_cln = j.fetchall()

              #for testing 2
              j = 0
              query_list2 = list()
              for j in range(len(parameters)):
                query = '''SELECT DISTINCT acc.`inptCaseAccInptCaseId`, acc.`inptCaseAccChqNum`
                          FROM inpt_case_accounting acc
                          WHERE acc.inptCaseAccType = "BD" AND acc.`inptCaseAccInptCaseId` = {0}'''.format(parameters[j])
                query_list2.append(query)
      
              i = 0
              results_cln = pd.DataFrame()
              for i in range(len(query_list2)):
                cur_mysql.execute(query_list2[i])
                result = cur_mysql.fetchall()
                results_cln = results_cln.append(result)

              #End of testing code


              fetched_result = fetched_result.drop_duplicates()
              results_cln = results_cln.drop_duplicates()
              fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg','CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                         'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName',  'CaseMmClientName',
                                         'CaseMmInsurerName', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate', 'CasePpId']
              results_cln.columns = ['Sub Case ID', 'Client Listing Number']
              if results_cln.empty != True:
                fetched_result['CaseAccChqNum'] = results_cln['Client Listing Number']
              else:
                fetched_result['CaseAccChqNum'] = "N/A"
            else:
              #stored_proc = cur_mysql.callproc('query_marc_for_inpatient_validity', parameters)
              #for z in cur_mysql.stored_results():
              #  results = z.fetchall()
              #for testing
              z = 0
              query_list = list()
              for z in range(len(parameters)):
                query = '''SELECT DISTINCT cse.`inptCaseId`, cse.`inptCaseB2B`, cse.`inptCaseSpecArrg`,cse.`inptCaseHospitalName`, cse.`inptCaseAdmDate`,
                inpt.`inptCaseMmExtRefId`, inpt.`inptCasemmPolicyNum`, inpt.`inptCaseMmPlanExtCode`, inpt.`inptCaseMmPrgSvcId`, inpt.`inptCasemmServiceName`, inpt.`inptCaseMmPrgId`, inpt.`inptCaseMmPrgName`, inpt.`inptCaseMmPlanName`, inpt.`inptCaseMmPolType`, inpt.`inptCaseMmNRIC`, inpt.`inptCaseMmName`,  inpt.`inptCaseMmClientName`, inpt.`inptCaseMmInsurerName`, acc.`inptCaseAccChqNum`, ob.`iactwsCompPayAmt`, ob.`iactwsBillNum`, inptcase.`inptCaseBillRegDate`, inptcase.`inptCaseBillReceivedDate`, inptcase.`inptCaseDiscDate`
                FROM inpt_case cse
                INNER JOIN `inpt_case_member` inpt ON inpt.`inptCaseMmId` = cse.`inptCaseMemberId`
                INNER JOIN `client` cli ON cli.`clientId` = inpt.`inptCaseMmInsurerId`
                INNER JOIN `inpt_case_accounting` acc ON acc.`inptCaseAccInptCaseId` = cse.`inptCaseId`
                INNER JOIN `inpt_case_worksheet_value` ob ON ob.`iactwsId` = cse.`inptCaseInptCaseWorksheetValueId`
                INNER JOIN `inpt_case` inptcase ON inptcase.`inptCaseMemberId` = cse.`inptCaseMemberId`
                WHERE acc.inptCaseAccType = "BD" AND cse.`inptCaseId` = {0}'''.format(parameters[z])
                query_list.append(query)

              i = 0
              results = list()
              fetched_result = pd.DataFrame()
              for i in range(len(query_list)):
                cur_mysql.execute(query_list[i])
                result = cur_mysql.fetchall()
                results.append(result)
                fetched_result = fetched_result.append(results[i])

              #End of testing code
              #fetched_result = pd.DataFrame(results)
              fetched_result = fetched_result.drop_duplicates()
              fetched_result.columns = ['CaseId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate', 'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode', 'CaseMmPrgSvcId',
                                         'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName', 'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName', 'CaseMmClientName',
                                         'CaseMmInsurerName', 'CaseAccChqNum', 'CompPayAmt', 'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate', 'inptCaseDiscDate']
              fetched_result['CasePpId'] = "0"

        # Rename Policy Type
        value = fetched_result['CaseMmPolType'][0]
        if value == 'GL':
          fetched_result['InsurerType'] = 'GROUP'
        elif value == 'IL':
          fetched_result['InsurerType'] = 'INDIVIDUAL'
        elif value == 'GG':
          fetched_result['InsurerType'] = 'GROUP'
        elif value == 'IG':
          fetched_result['InsurerType'] = 'INDIVIDUAL'

        if document == 'REQ':
          fetched_result['InsurerType'] = 'FLOWER'
        else:
          fetched_result['InsurerType'] = fetched_result['InsurerType']

        # Search for the Insurer's address and issuer's name
        extracted_row = merged_insurer_details.loc[merged_insurer_details['CaseMmPolicyNum'] == '%s' % str(fetched_result['CasemmPolicyNum'][0])]
        extracted_row = extracted_row.loc[extracted_row['CaseMmPlanExtCode'] == '%s' % fetched_result['CaseMmPlanExtCode'][0]]
        extracted_row = extracted_row.loc[extracted_row['CaseMmPrgName'] == '%s' % fetched_result['CaseMmPrgName'][0]]
        extracted_row = extracted_row.loc[extracted_row['CaseMmPolType'] == '%s' % fetched_result['CaseMmPolType'][0]]
        extracted_row = extracted_row.loc[extracted_row['CaseMmPlanName'] == '%s' % fetched_result['CaseMmPlanName'][0]]
        extracted_row = extracted_row.reset_index()
        del extracted_row['index']
        extracted_row = extracted_row.drop_duplicates()

        # Get the correct address
        rownum = extracted_row['CaseMmPlanName'].count()
        if rownum == 1:
          fetched_result['Attention To'] = extracted_row['atten_1']
          fetched_result['Address'] = extracted_row['comp_address']
          fetched_result['client_id'] = extracted_row['id']
        elif rownum > 1:
          extracted_row = extracted_row.loc[extracted_row['identifier'] == '%s' % fetched_result['InsurerType'][0]]
          extracted_row = extracted_row.drop_duplicates()
          extracted_row = extracted_row.set_index('CaseMmPlanName')
          extracted_row = extracted_row.reset_index()
          rownum2 = extracted_row['CaseMmPlanName'].count()
          if rownum2 == 1:
            fetched_result['Attention To'] = extracted_row['atten_1']
            fetched_result['Address'] = extracted_row['comp_address']
            fetched_result['client_id'] = extracted_row['id']
          elif rownum2 > 1:
            fetched_result['Attention To'] = "Too many possible information. Please manually check."
            fetched_result['Address'] = "Too many possible information. Please manually check."
            fetched_result['client_id'] = 0
          elif rownum2 == 0:
            fetched_result['Attention To'] = "No possible information. Please manually check."
            fetched_result['Address'] = "No possible information. Please manually check."
            fetched_result['client_id'] = 0
        elif rownum > 1:
          fetched_result['Attention To'] = "Too many possible information. Please manually check."
          fetched_result['Address'] = "Too possible information. Please manually check."
          fetched_result['client_id'] = 0
        else:
          fetched_result['Attention To'] = "No possible information. Please manually check."
          fetched_result['Address'] = "No possible information. Please manually check."
          fetched_result['client_id'] = 0

        # Insert temporary H&S running number
        fetched_result['HNS Number'] = str("%s") % HNS
        newHNS = HNS.split("-")
        a = 0
        b = 1

        if document == 'REQ':
          HNS = newHNS[a] + "-" + str(int(newHNS[b]) + 1).zfill(3)
        else:
          HNS = newHNS[a] + "-" + str(int(newHNS[b]) + 1).zfill(5)

        if type == 'POST':
          myjason = pd.DataFrame(jsondf_valid.loc[jsondf_valid['Case ID'] == int(fetched_result['CasePpId'])])
          myjason = pd.DataFrame(myjason.loc[myjason['Amount'] == jsondf_valid['Amount'][i]])
          myjason['Case ID'] = pd.to_numeric(myjason['Case ID'])
          fetched_result = pd.merge(left = fetched_result, right = myjason, left_on = fetched_result['CasePpId'], right_on = myjason['Case ID'])
        else:
          myjason = pd.DataFrame(jsondf_valid.loc[jsondf_valid['Case ID'] == int(fetched_result['CaseId'][0])])
          myjason = pd.DataFrame(myjason.loc[myjason['Amount'] == jsondf_valid['Amount'][i]])
          myjason['Case ID'] = pd.to_numeric(myjason['Case ID'])
          fetched_result = pd.merge(left = fetched_result, right = myjason, left_on = fetched_result['CaseId'], right_on = myjason['Case ID'])
          fetched_result = fetched_result.drop_duplicates()

        # Consolidated dataframes
        temp_df = temp_df.append(fetched_result, ignore_index = True, sort = False)
        temp_df = temp_df.drop_duplicates()

      # Finalizing valid Case IDs
      temp_df = temp_df.drop(columns = ['key_0', 'Case ID'])
      temp_df = temp_df[['CaseId', 'CasePpId', 'CaseB2B', 'CaseSpecArrg', 'CaseHospitalName', 'CaseAdmDate',
       'CaseMmExtRefId', 'CasemmPolicyNum', 'CaseMmPlanExtCode',
       'CaseMmPrgSvcId', 'CasemmServiceName', 'CaseMmPrgId', 'CaseMmPrgName',
       'CaseMmPlanName', 'CaseMmPolType', 'CaseMmNRIC', 'CaseMmName',
       'CaseMmClientName', 'CaseMmInsurerName', 'CaseAccChqNum', 'CompPayAmt',
       'BillNum', 'inptCaseBillRegDate', 'inptCaseBillReceivedDate',
       'inptCaseDiscDate', 'InsurerType', 'Attention To',
       'Address', 'client_id', 'HNS Number', 'Amount', 'Reason', 'Remarks',
       'Type', 'Validity', 'Client', 'Arrangement']]

      CaseAccChqNum_column = pd.DataFrame({'Client ID' : temp_df['client_id'].to_list(), 'CaseAccChqNum': temp_df['CaseAccChqNum'].to_list()})
      CaseAccChqNum_column = CaseAccChqNum_column.fillna('None')

      # Rename these columns
      temp_df.columns = ['Case ID', 'Sub Case ID', 'B2B', 'Special Arrangement', 'Hospital Name', 'Admission Date', 'Member Ref. ID', 'Insurence Policy Number', 'Plan Ext. Code',
                         'Program Service ID', 'Service Name', 'Program ID', 'Program Name', 'Insurence Plan Name', 'Policy Type', 'IC Number', 'Patient Name', 'Client Name',
                         'Insurer Name', 'Client Listing Number', 'Company Payment Amount', 'Bill. Num.', 'OB Registered Date', 'OB Received Date', 'Discharged Date', 'Insurer Type',
                         'Attention To', 'Address', 'Client ID', 'HNS Number', 'Amount', 'Reason', 'Remarks', 'Type', 'Validity', 'Client', 'Arrangement']

      # ONLY FOR REQ
      if str(document) == 'REQ' and str(temp_df['Insurer Name'][i]) != 'Allianz General Insurance Company (M) Bhd':
        temp_df['Client Listing Number'] = 'N/A'
      else:
        temp_df['Client Listing Number'] = temp_df['Client Listing Number']

      # Process the invalid Case IDs
      jsondf_invalid = jsondf_val.loc[jsondf_val['Validity'] != "Valid Case ID"].set_index('Case ID').reset_index()
      v = 0
      invalidid = jsondf_invalid['Case ID']
      for i in range(len(invalidid)):
        invalid_df.loc[i, 'Case ID'] = "Case ID [ %s ] cannot be found in MARC. Please check." % invalidid[i]
        invalid_df.loc[i, 'Amount'] = int(jsondf_invalid['Amount'][i])
        invalid_df.loc[i, 'Remarks'] = jsondf_invalid['Remarks'][i]
        invalid_df.loc[i, 'Type'] = jsondf_invalid['Type'][i]
        invalid_df.loc[i, 'Reason'] = jsondf_invalid['Reason'][i]
        invalid_df.loc[i, 'Validity'] = jsondf_invalid['Validity'][i]

      # Consolidating temp_df with invalid_df to get a finalized dataframe
      temp_df = temp_df.append(invalid_df, ignore_index = True, sort = False)
      #temp_df = temp_df.merge(CaseAccChqNum_column, how = 'left', on = 'Client ID')
      temp_df = temp_df.drop_duplicates()
      temp_df = pd.DataFrame(temp_df)

      # close the communication with the MARC database
      cur_mysql.close()

    # Return the dataframe to workflow script
    return temp_df
  except Exception as error:
    print(error)

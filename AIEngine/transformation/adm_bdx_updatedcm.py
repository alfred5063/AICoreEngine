#!/usr/bin/python
# FINAL SCRIPT updated as of 8th July 2020
# Workflow - CBA/ADMISSION

# Declare Python libraries needed for this script
import openpyxl
from datetime import datetime as dt
import ast
import pandas as pd
from xlrd import *
from collections import OrderedDict
from connector.connector import MySqlConnector,MSSqlConnector
from connector.dbconfig import *
import pyodbc as db
import psycopg2
from loading.query.query_as_df import *
from transformation.adm_bdx_manipulation import count_BdxL,read_BdxL
import json
import datetime as d
import time
from loading.excel.checkExcelLoading import is_locked
from utils.audit_trail import audit_log
from datetime import datetime
from transformation.adm_connect_db import *
from directory.get_filename_from_path import *
from directory.createfolder import *
from transformation.reqdca_excel_manipulation import *


def get_first_case_id(file):
  rb=open_workbook(file,formatting_info=True)
  r_sheet=rb.sheet_by_index(0)
  ID=str(r_sheet.cell(17, 1).value)
  if len(ID)>6:
    Case_ID_List = ID.split("-")
    ID=Case_ID_List[1]
  return ID

def Check_post_or_type(caseid):
  conn = MySqlConnector()
  cur = conn.cursor()

  # Calling Stored Procedure
  parameters = [str(caseid)]
  stored_proc = cur.callproc("adm_query_marc_by_caseid",parameters)
  for i in cur.stored_results():
    results = i.fetchall()
  fetched_result = pd.DataFrame(results)
  
  cur.close()
  conn.close()

  if fetched_result.empty == True:
    print("NOT POST")
    type="ADM"
  else:
    print("POST")
    type="POST"
  return type

def get_Bdx_Total_Amount(file):
  rb=open_workbook(file,formatting_info=True)
  readerSheet = rb.sheet_by_index(0)
  for i in range(20):
    if "No" == str(readerSheet.cell(i,0).value):
      skip_row=i
  df=read_BdxL(file).T
  total=0
  for i in df.loc["Insurance"]:
    total+=i
  output=round(total,2)
  column=0
  for i in (df.index):
    if i == "Insurance":
      break
    column+=1
  row=skip_row+count_BdxL(file)+8
  readerSheet = open_workbook(file).sheet_by_index(0)
  thisCell = readerSheet.cell(row, column).value
  return thisCell

def dis_claim_no_to_index(dis_claim_no):
  temp=dis_claim_no.split('-')
  return int(temp[1])+4


def bdx_to_DCM(BDFile,DCMFile,invoice_type,adm_base,path=None):
  if path==None:
    path=DCMFile


  if is_locked(DCMFile):
    audit_log("DCMFILE cannot be opened", "Completed...", adm_base)
    return None
  else:
    audit_log("DCMFILE can be opened", "Completed...", adm_base)
    wb=openpyxl.load_workbook(DCMFile)

  rb=open_workbook(BDFile,formatting_info=True)
  readerSheet = rb.sheet_by_index(0)
  for i in range(20):
    if "Bord No" in str(readerSheet.cell(i,0).value):
      bdx_lisitng=readerSheet.cell(i,1).value
    if "Submission" in str(readerSheet.cell(i,0).value):
      date=readerSheet.cell(i,1).value
    if "Insurance" in str(readerSheet.cell(i,0).value):
      insurance_name=readerSheet.cell(i,1).value

  df=read_BdxL(BDFile)
  Disbursement_No=df["Disbursement No"][0]
  row=get_unfill_row(DCMFile)
  #row=dis_claim_no_to_index(Disbursement_No)
  no_of_case=count_BdxL(BDFile)

  readerSheet_ack=rb.sheet_by_index(0)
  Claims_listing_no=readerSheet_ack.cell(12,2).value
  bdx_lisitng=bdx_lisitng.replace(Claims_listing_no,"")
  ws=wb["Sheet1"]
  ws.cell(row=row, column=1).value=date
  ws.cell(row=row, column=2).value="Disbursement claims"
  ws.cell(row=row, column=3).value=Disbursement_No
  ws.cell(row=row, column=5).value=no_of_case
  ws.cell(row=row, column=7).value=insurance_name
  ws.cell(row=row, column=4).value=bdx_lisitng
  ws.cell(row=row, column=7).value=insurance_name
  ws.cell(row=row, column=11).value=get_Bdx_Total_Amount(BDFile)
  ws.cell(row=row, column=12).value=Check_post_or_type(get_first_case_id(BDFile))
  if is_locked(path):
    audit_log("DCM file is locked,data is insert into DB and dummy file", "Completed...", adm_base)
  else:
    wb.save(path)
    audit_log("DCMFILE is saved", "Completed...", adm_base)
    print("DCMFILE is saved")


def get_unfill_row(file):
  rd=open_workbook(file)
  r_sheet=rd.sheet_by_index(0)
  skiprow=0
  for i in range(20):
    if "date" in str(r_sheet.cell(i,0).value).lower():
      skiprow=i
      break
  data=pd.read_excel(file,skiprows = skiprow , na_values = "Missing")
  last_column=0
  colname = data.columns[0]
  col_list=pd.isna(data[colname])
  print(col_list)
  for i in range(len(col_list)):
    if col_list[i] == True and col_list[i+1] == True:
      print(col_list[i])
      last_column=i
      break
  return last_column+1+skiprow+1


def bdx_to_DCM_database(BDFile, invoice_type, adm_base):

  conn_mysql = MySqlConnector()
  cur_mysql = conn_mysql.cursor()

  email = adm_base.email
  rb = open_workbook(BDFile,formatting_info=True)
  readerSheet = rb.sheet_by_index(0)
  for i in range(20):
    if "Bord No" in str(readerSheet.cell(i,0).value):
      bdx_lisitng=readerSheet.cell(i,1).value
    if "Submission" in str(readerSheet.cell(i,0).value):
      date=readerSheet.cell(i,1).value
    if "Insurance" in str(readerSheet.cell(i,0).value):
      insurance_name=readerSheet.cell(i,1).value

  bdx_lisitng = bdx_lisitng.split("/")
  bdx_lisitng = str(bdx_lisitng[0].zfill(5)) + "/" + str(bdx_lisitng[1])

  df = read_BdxL(BDFile)
  Disbursement_No=df["Disbursement No"][0]
  no_of_case=count_BdxL(BDFile)
  dm_json = { "Date":"", "Types of Invoice":"", "Disbursement Claims No":"", "Claims Listing No":"", "No of Cases":"", "File No":"", "Customer Name Master":"",
            "Hospital":"", "Bill No":"", "Patient":"", "Bill Amount":"", "Reasons":"", "Action":"", "Initial":"", "Bank in Date":"", "Cheque No or TT":"",
            "Cheque Amount Paid":"", "Remarks":"", "Reason for Adjustment":"", "Next Action Taken":"", "New invoice no":"", "Invoice Amount ":"", "Issue Date":"",
            "Cheque No":"", "Settlement Amount":"", "OCBC 4":"", "DB 04":"", "DB 10":"", "DB 20":"", "DB 21":"", "DB 39":"", "BIMB 1":"", "DB 03":"", "DB 27":"",
            "HSBC 2":"", "AP Team Remarks":""}

  reason = Check_post_or_type(get_first_case_id(BDFile))
  total = get_Bdx_Total_Amount(BDFile)

  paramater = [str(email),]
  stored_proc = cur_mysql.callproc('dca_query_marc_for_user_email', paramater)
  for i in cur_mysql.stored_results():
    results = i.fetchall()
  if results != []:
    shortname = results[0][0]
  else:
    shortname = ''

  dm_json["Date"] = str(date).replace('/', '-')
  dm_json["Types of Invoice"] = "Disbursement claims"
  dm_json["Disbursement Claims No"] = str(Disbursement_No)
  dm_json["Claims Listing No"] = str(bdx_lisitng)
  dm_json["No of Cases"] = no_of_case
  dm_json["File No"] = ""
  dm_json["Customer Name Master"] = str(insurance_name)
  dm_json["Hospital"] = ""
  dm_json["Bill No"] = ""
  dm_json["Patient"] = ""
  dm_json["Bill Amount"] = str(total)
  dm_json["Reasons"] = str(reason)
  dm_json["Action"] = ""
  dm_json["Initial"] = shortname
  dm_json["Bank in Date"] = ""
  dm_json["Cheque No or TT"] = ""
  dm_json["Cheque Amount Paid"] = ""
  dm_json["Remarks"] = ""
  dm_json["Reason for Adjustment"] = ""
  dm_json["Next Action Taken"] = ""
  dm_json["New invoice no"] = ""
  dm_json["Invoice Amount"] = ""
  dm_json["Issue Date"] = ""
  dm_json["Cheque No"] = ""
  dm_json["Settlement Amount"] = ""
  dm_json["OCBC 4"] = ""
  dm_json["DB 04"] = ""
  dm_json["DB 10"] = ""
  dm_json["DB 20"] = ""
  dm_json["DB 21"] = ""
  dm_json["DB 39"] = ""
  dm_json["BIMB 1"] = ""
  dm_json["DB 03"] = ""
  dm_json["DB 27"] = ""
  dm_json["HSBC 2"] = ""
  dm_json["AP Team Remarks"] = ""

  #dcm_path = r'\\dtisvr2\CBA_UAT\Result\REQDCA_Paths\1. DISBURSEMENT CLAIMS INVOICE LOG FILE'
  dcm_path = r'\\fs1\finance\1-CREDIT CONTROL\2-MASTER & OPS INVOICES\1. DISBURSEMENT CLAIMS INVOICE LOG FILE'
  DCMfile = dcm_path + "\%s" % str(dt.now().year) + "\Disbursement Claims Master %s.xlsx" % str(dt.now().year)
  filename = get_file_name(DCMfile)
  if(os.path.isdir(str(filename[0]) + "\\TEMPORARY\\") == False):
    createfolder(str(filename[0]) + "\\TEMPORARY\\", adm_base, parents = True, exist_ok = True)
  if(os.path.isdir(str(filename[0]) + "\\BACKUP\\") == False):
    createfolder(str(filename[0]) + "\\BACKUP\\", adm_base, parents = True, exist_ok = True)
  myfile = list(getListOfFiles(str(filename[0]) + "\\TEMPORARY\\"))
  mylist = pd.DataFrame(list(filter(lambda x: str(filename[1]) in x, myfile)), columns = ['Filenames'])
  if mylist.empty == True:
    shutil.copy(DCMfile, str(filename[0]) + "\\BACKUP\\")
    shutil.move(DCMfile, str(filename[0]) + "\\TEMPORARY\\")
    myfile = list(getListOfFiles(str(filename[0]) + "\\TEMPORARY\\"))
    mylist = pd.DataFrame(list(filter(lambda x: str(filename[1]) in x, myfile)), columns = ['Filenames'])
    filelocation = mylist['Filenames'][0]
  else:
    filelocation = mylist['Filenames'][0]

  # Load Disbursement Claim Master File
  wb = load_workbook(filelocation)
  sheetname = wb.sheetnames
  sheet = wb.index(wb.active)
  activesheet = wb.active
  start_cell = search_excel('Date', sheet, wb)
  end_cell = search_excel('Reason for adjustment', sheet, wb)
  last_row = search_last_row(start_cell, filelocation, sheet, wb, 'Date')
  start_row = last_row + 1

  # Create empty dataframe with header
  dcm_df = pd.DataFrame(pd.read_excel(filelocation, sheet_name = sheet, skiprows = start_cell[1]-1, nrows = 0))
  dcm_df = dcm_df.rename(columns = {"Types of invoice ": "Types of invoice"})
  dcm_df.loc[0, 'Date'] = str(date).replace('/', '-')
  dcm_df.loc[0,'DCA'] = 'Disbursement claims'
  dcm_df.loc[0,'Dis. claims no'] = df["Disbursement No"][0]
  dcm_df.loc[0,'No. of cases'] = no_of_case
  dcm_df.loc[0,'File no'] = df["File No"][0]
  dcm_df.loc[0,'Claims listing no'] = str(bdx_lisitng)
  dcm_df.loc[0,'Customer Name_Master'] = str(insurance_name)
  dcm_df.loc[0,'Amounts'] = str(total)
  dcm_df.loc[0,'Reasons'] = str(reason)
  dcm_df.loc[0,'Initial'] = shortname

  update_entry(activesheet, dcm_df, start_row, 0, start_column = 1)
  wb.save(str(DCMfile))
  os.remove(filelocation)

  try:
    update_dcm_to_database(dm_json,email)
  except:
    update_new_dcm_to_database(dm_json,email)


